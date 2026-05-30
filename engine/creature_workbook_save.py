from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, timedelta
import os
from pathlib import Path
import re
import shutil
import uuid

from openpyxl import load_workbook

from engine.excel_creatures import (
    ACTION_COLUMNS,
    CREATURES_SHEET,
    SIM_COVERAGE_ERROR,
    action_card_coverage,
    build_creature_action_card,
)


class CreatureWorkbookSaveError(ValueError):
    pass


STAT_COLUMN_MAP = {
    "power": "Power",
    "toughness": "Toughness",
    "magicArmor": "Magic_Armor",
    "armor": "Armor",
    "baseGuard": "Base_Guard",
    "movement": "Move",
    "threatLevel": "Threat_Level",
}

STAT_MINIMUMS = {
    "power": 1,
    "toughness": 1,
    "magicArmor": 0,
    "armor": 0,
    "baseGuard": 0,
    "movement": 1,
    "threatLevel": 0,
}

SKILL_COLUMN_MAP = {
    "intelligence": "Intelligence",
    "alertness": "Alertness",
    "stealth": "Stealth",
    "social": "Social",
    "arcana": "Arcana",
    "athletics": "Athletics",
}

ACTION_RESULTS = set(ACTION_COLUMNS)
BACKUP_DIR_NAME = "creature_workbook_backups"
BACKUP_TIMESTAMP_RE = re.compile(r"^(.+)__(\d{8}_\d{6})(?:_\d+)?\.xlsx$", flags=re.I)


@dataclass(frozen=True)
class BackupEntry:
    path: Path
    timestamp: datetime


def save_creature_overrides_to_workbook(
    *,
    workbook_path: Path,
    template_id: str,
    overrides: dict,
    backup_dir: Path,
    now: datetime | None = None,
) -> dict:
    now = now or datetime.now()
    workbook_path = Path(workbook_path)
    backup_dir = Path(backup_dir)
    updates = _normalize_updates(overrides)
    if not updates:
        raise CreatureWorkbookSaveError("No overrides to save")

    try:
        workbook = load_workbook(workbook_path)
    except PermissionError as exc:
        raise CreatureWorkbookSaveError("Could not read creature workbook; close the workbook and retry") from exc
    except OSError as exc:
        raise CreatureWorkbookSaveError(f"Could not read creature workbook: {exc}") from exc

    temp_path: Path | None = None
    try:
        sheet = _worksheet(workbook)
        headers = _headers(sheet)
        row_index = _row_for_template(sheet, headers, template_id)
        _validate_columns(headers, updates)

        backup_path = _create_backup(workbook_path, backup_dir, now=now)
        for column_name, value in updates.items():
            sheet.cell(row=row_index, column=headers[column_name]).value = value

        temp_path = workbook_path.with_name(f".{workbook_path.stem}.{uuid.uuid4().hex}.tmp.xlsx")
        workbook.save(temp_path)
        os.replace(temp_path, workbook_path)
        temp_path = None
    except PermissionError as exc:
        raise CreatureWorkbookSaveError("Could not write creature workbook; close the workbook and retry") from exc
    except OSError as exc:
        raise CreatureWorkbookSaveError(f"Could not write creature workbook: {exc}") from exc
    finally:
        workbook.close()
        if temp_path is not None and temp_path.exists():
            temp_path.unlink(missing_ok=True)

    deleted = prune_workbook_backups(backup_dir, workbook_path.name, now=now)
    return {
        "backupFilename": backup_path.name,
        "backupPath": str(backup_path),
        "updatedColumns": sorted(updates),
        "deletedBackups": [path.name for path in deleted],
    }


def prune_workbook_backups(
    backup_dir: Path,
    workbook_name: str,
    *,
    now: datetime | None = None,
) -> list[Path]:
    now = now or datetime.now()
    entries = _backup_entries(Path(backup_dir), workbook_name)
    keep: set[Path] = set()

    recent: list[BackupEntry] = []
    daily: dict[object, BackupEntry] = {}
    weekly: dict[object, BackupEntry] = {}
    monthly: dict[object, BackupEntry] = {}

    for entry in entries:
        age = now - entry.timestamp
        if age < timedelta(days=1):
            recent.append(entry)
        elif age < timedelta(days=7):
            daily.setdefault(entry.timestamp.date(), entry)
        elif age < timedelta(days=31):
            iso = entry.timestamp.isocalendar()
            weekly.setdefault((iso.year, iso.week), entry)
        else:
            monthly.setdefault((entry.timestamp.year, entry.timestamp.month), entry)

    keep.update(entry.path for entry in recent[:25])
    keep.update(entry.path for entry in daily.values())
    keep.update(entry.path for entry in weekly.values())
    keep.update(entry.path for entry in monthly.values())

    deleted: list[Path] = []
    for entry in entries:
        if entry.path in keep:
            continue
        entry.path.unlink(missing_ok=True)
        deleted.append(entry.path)
    return deleted


def _normalize_updates(overrides: dict | None) -> dict[str, int | str]:
    if not isinstance(overrides, dict):
        raise CreatureWorkbookSaveError("Overrides must be an object")
    updates: dict[str, int | str] = {}
    updates.update(_normalize_stat_updates(overrides.get("statOverrides") or {}))
    updates.update(_normalize_skill_updates(overrides.get("skillOverrides") or {}))
    updates.update(_normalize_action_updates(overrides.get("actionOverrides") or {}))
    return updates


def _normalize_stat_updates(raw: object) -> dict[str, int]:
    if not raw:
        return {}
    if not isinstance(raw, dict):
        raise CreatureWorkbookSaveError("statOverrides must be an object")
    updates: dict[str, int] = {}
    for key, value in raw.items():
        if value is None or value == "":
            continue
        if key == "initiativeModifier":
            raise CreatureWorkbookSaveError("initiativeModifier is derived from Alertness; edit Alertness instead")
        if key not in STAT_COLUMN_MAP:
            raise CreatureWorkbookSaveError(f"Unknown stat override '{key}'")
        number = _int_value(value, f"{key} override")
        minimum = STAT_MINIMUMS[key]
        if number < minimum:
            raise CreatureWorkbookSaveError(f"{key} override must be >= {minimum}")
        updates[STAT_COLUMN_MAP[key]] = number
    return updates


def _normalize_skill_updates(raw: object) -> dict[str, int]:
    if not raw:
        return {}
    if not isinstance(raw, dict):
        raise CreatureWorkbookSaveError("skillOverrides must be an object")
    updates: dict[str, int] = {}
    for key, value in raw.items():
        if value is None or value == "":
            continue
        if key not in SKILL_COLUMN_MAP:
            raise CreatureWorkbookSaveError(f"Unknown skill override '{key}'")
        number = _int_value(value, f"{key} skill override")
        if number < 0:
            raise CreatureWorkbookSaveError(f"{key} skill override must be >= 0")
        updates[SKILL_COLUMN_MAP[key]] = number
    return updates


def _normalize_action_updates(raw: object) -> dict[str, str]:
    if not raw:
        return {}
    if not isinstance(raw, dict):
        raise CreatureWorkbookSaveError("actionOverrides must be an object")
    updates: dict[str, str] = {}
    for raw_result, raw_text in raw.items():
        result = str(raw_result or "").strip().upper()
        if result not in ACTION_RESULTS:
            raise CreatureWorkbookSaveError(f"Unknown action override '{raw_result}'")
        text = str(raw_text or "").strip()
        if not text:
            raise CreatureWorkbookSaveError(f"{result} action override text is required")
        card = build_creature_action_card(
            creature_id="__save_preview__",
            action_result=result,
            action_text=text,
        )
        coverage = action_card_coverage(card)
        if coverage["status"] == SIM_COVERAGE_ERROR:
            raise CreatureWorkbookSaveError(f"{result} action override is not simulatable")
        updates[result] = text
    return updates


def _int_value(value: object, label: str) -> int:
    try:
        return int(value)
    except (TypeError, ValueError) as exc:
        raise CreatureWorkbookSaveError(f"{label} must be an integer") from exc


def _worksheet(workbook):
    if CREATURES_SHEET not in workbook.sheetnames:
        raise CreatureWorkbookSaveError(f"Sheet '{CREATURES_SHEET}' not found")
    return workbook[CREATURES_SHEET]


def _headers(sheet) -> dict[str, int]:
    headers: dict[str, int] = {}
    for cell in sheet[1]:
        header = str(cell.value or "").strip()
        if header:
            headers[header] = int(cell.column)
    if "ID" not in headers:
        raise CreatureWorkbookSaveError("Creature workbook is missing ID column")
    return headers


def _row_for_template(sheet, headers: dict[str, int], template_id: str) -> int:
    wanted = str(template_id or "").strip()
    if not wanted:
        raise CreatureWorkbookSaveError("templateId is required")
    id_column = headers["ID"]
    for row_index in range(2, sheet.max_row + 1):
        current = str(sheet.cell(row=row_index, column=id_column).value or "").strip()
        if current == wanted:
            return row_index
    raise CreatureWorkbookSaveError(f"Unknown template '{wanted}'")


def _validate_columns(headers: dict[str, int], updates: dict[str, int | str]) -> None:
    missing = sorted(column for column in updates if column not in headers)
    if missing:
        raise CreatureWorkbookSaveError(f"Creature workbook is missing column '{missing[0]}'")


def _create_backup(workbook_path: Path, backup_dir: Path, *, now: datetime) -> Path:
    backup_dir.mkdir(parents=True, exist_ok=True)
    stem = workbook_path.stem
    timestamp = now.strftime("%Y%m%d_%H%M%S")
    target = backup_dir / f"{stem}__{timestamp}.xlsx"
    suffix = 2
    while target.exists():
        target = backup_dir / f"{stem}__{timestamp}_{suffix}.xlsx"
        suffix += 1
    try:
        shutil.copy2(workbook_path, target)
    except PermissionError as exc:
        raise CreatureWorkbookSaveError("Could not back up creature workbook; close the workbook and retry") from exc
    except OSError as exc:
        raise CreatureWorkbookSaveError(f"Could not back up creature workbook: {exc}") from exc
    return target


def _backup_entries(backup_dir: Path, workbook_name: str) -> list[BackupEntry]:
    if not backup_dir.exists():
        return []
    stem = Path(workbook_name).stem
    entries: list[BackupEntry] = []
    for path in backup_dir.glob(f"{stem}__*.xlsx"):
        match = BACKUP_TIMESTAMP_RE.match(path.name)
        if not match or match.group(1) != stem:
            continue
        try:
            timestamp = datetime.strptime(match.group(2), "%Y%m%d_%H%M%S")
        except ValueError:
            continue
        entries.append(BackupEntry(path=path, timestamp=timestamp))
    return sorted(entries, key=lambda entry: (entry.timestamp, entry.path.name), reverse=True)
