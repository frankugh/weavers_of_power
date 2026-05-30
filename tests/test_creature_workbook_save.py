from __future__ import annotations

from datetime import datetime, timedelta
import shutil
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from engine.creature_workbook_save import prune_workbook_backups, save_creature_overrides_to_workbook
from engine.excel_creatures import CREATURES_SHEET


PROJECT_ROOT = Path(__file__).resolve().parents[1]


def workbook_cell_values(workbook_path: Path, template_id: str, columns: list[str]) -> dict[str, object]:
    workbook = load_workbook(workbook_path)
    try:
        sheet = workbook[CREATURES_SHEET]
        headers = {str(cell.value).strip(): int(cell.column) for cell in sheet[1] if cell.value}
        id_column = headers["ID"]
        for row_index in range(2, sheet.max_row + 1):
            if str(sheet.cell(row=row_index, column=id_column).value or "").strip() != template_id:
                continue
            return {
                column: sheet.cell(row=row_index, column=headers[column]).value
                for column in columns
            }
    finally:
        workbook.close()
    raise AssertionError(f"Template {template_id} not found")


class CreatureWorkbookSaveTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = tempfile.TemporaryDirectory()
        self.addCleanup(self.temp_dir.cleanup)
        self.workbook_path = Path(self.temp_dir.name) / "denizens_creature_database.xlsx"
        shutil.copy2(PROJECT_ROOT / "data" / "denizens_creature_database.xlsx", self.workbook_path)
        self.backup_dir = Path(self.temp_dir.name) / "backups"

    def test_save_updates_stats_skills_and_actions_with_backup(self) -> None:
        result = save_creature_overrides_to_workbook(
            workbook_path=self.workbook_path,
            template_id="C_GOBLIN",
            overrides={
                "statOverrides": {"toughness": 13, "armor": 2, "movement": 7},
                "skillOverrides": {"alertness": 6, "stealth": 5},
                "actionOverrides": {"A1": "Excel Strike - Attack 9 pierce 2"},
            },
            backup_dir=self.backup_dir,
            now=datetime(2026, 5, 30, 12, 0, 0),
        )

        self.assertEqual(result["backupFilename"], "denizens_creature_database__20260530_120000.xlsx")
        self.assertTrue((self.backup_dir / result["backupFilename"]).exists())
        values = workbook_cell_values(
            self.workbook_path,
            "C_GOBLIN",
            ["Toughness", "Armor", "Move", "Alertness", "Stealth", "A1"],
        )
        self.assertEqual(values["Toughness"], 13)
        self.assertEqual(values["Armor"], 2)
        self.assertEqual(values["Move"], 7)
        self.assertEqual(values["Alertness"], 6)
        self.assertEqual(values["Stealth"], 5)
        self.assertEqual(values["A1"], "Excel Strike - Attack 9 pierce 2")

    def test_invalid_save_inputs_raise_clear_errors(self) -> None:
        with self.assertRaisesRegex(ValueError, "Alertness"):
            save_creature_overrides_to_workbook(
                workbook_path=self.workbook_path,
                template_id="C_GOBLIN",
                overrides={"statOverrides": {"initiativeModifier": 4}},
                backup_dir=self.backup_dir,
            )

        with self.assertRaisesRegex(ValueError, "not simulatable"):
            save_creature_overrides_to_workbook(
                workbook_path=self.workbook_path,
                template_id="C_GOBLIN",
                overrides={"actionOverrides": {"A1": "Broken - Attack target"}},
                backup_dir=self.backup_dir,
            )

        with self.assertRaisesRegex(ValueError, "unknownStat"):
            save_creature_overrides_to_workbook(
                workbook_path=self.workbook_path,
                template_id="C_GOBLIN",
                overrides={"statOverrides": {"unknownStat": 1}},
                backup_dir=self.backup_dir,
            )

    def test_backup_retention_keeps_recent_daily_weekly_and_monthly(self) -> None:
        now = datetime(2026, 5, 30, 12, 0, 0)
        self.backup_dir.mkdir(parents=True)

        def touch_backup(stamp: datetime, suffix: str = "") -> Path:
            path = self.backup_dir / f"denizens_creature_database__{stamp.strftime('%Y%m%d_%H%M%S')}{suffix}.xlsx"
            path.write_text("backup", encoding="utf-8")
            return path

        recent_paths = [touch_backup(now - timedelta(minutes=minutes)) for minutes in range(30)]
        daily_old = touch_backup(now - timedelta(days=2, hours=3))
        daily_new = touch_backup(now - timedelta(days=2, hours=1), "_2")
        weekly_old = touch_backup(now - timedelta(days=12))
        weekly_new = touch_backup(now - timedelta(days=10))
        monthly_old = touch_backup(now - timedelta(days=75))
        monthly_new = touch_backup(now - timedelta(days=70))

        deleted = prune_workbook_backups(self.backup_dir, "denizens_creature_database.xlsx", now=now)

        remaining = {path.name for path in self.backup_dir.glob("*.xlsx")}
        self.assertEqual(sum(1 for path in recent_paths if path.name in remaining), 25)
        self.assertNotIn(daily_old.name, remaining)
        self.assertIn(daily_new.name, remaining)
        self.assertNotIn(weekly_old.name, remaining)
        self.assertIn(weekly_new.name, remaining)
        self.assertNotIn(monthly_old.name, remaining)
        self.assertIn(monthly_new.name, remaining)
        self.assertGreaterEqual(len(deleted), 8)


if __name__ == "__main__":
    unittest.main()
