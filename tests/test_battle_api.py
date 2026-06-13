from __future__ import annotations

import shutil
import tempfile
import unittest
from pathlib import Path

from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from battle_api import register_battle_api
from battle_session import BattleSessionContext
from engine.combat import WOUND_CARD_ID
from engine.models import Card, Effect

PROJECT_ROOT = Path(__file__).resolve().parents[1]


class BattleApiTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = tempfile.TemporaryDirectory()
        self.addCleanup(self.temp_dir.cleanup)
        saves_dir = Path(self.temp_dir.name) / "saves"
        self.context = BattleSessionContext(
            root=PROJECT_ROOT,
            saves_dir=saves_dir,
            map_templates_dir=Path(self.temp_dir.name) / "map_templates",
            scenarios_dir=Path(self.temp_dir.name) / "scenarios",
        )
        app = FastAPI()
        register_battle_api(app, self.context)
        self.client = TestClient(app)

    def _client_with_temp_workbook(self) -> tuple[TestClient, BattleSessionContext, Path]:
        workbook_path = Path(self.temp_dir.name) / "denizens_creature_database.xlsx"
        shutil.copy2(PROJECT_ROOT / "data" / "denizens_creature_database.xlsx", workbook_path)
        saves_dir = Path(self.temp_dir.name) / "workbook_saves"
        context = BattleSessionContext(
            root=PROJECT_ROOT,
            saves_dir=saves_dir,
            creatures_workbook=workbook_path,
            map_templates_dir=Path(self.temp_dir.name) / "workbook_map_templates",
            scenarios_dir=Path(self.temp_dir.name) / "workbook_scenarios",
        )
        app = FastAPI()
        register_battle_api(app, context)
        return TestClient(app), context, workbook_path

    def _workbook_values(self, workbook_path: Path, template_id: str, columns: list[str]) -> dict[str, object]:
        workbook = load_workbook(workbook_path)
        try:
            sheet = workbook["Creatures_Master"]
            headers = {str(cell.value).strip(): int(cell.column) for cell in sheet[1] if cell.value}
            id_column = headers["ID"]
            for row_index in range(2, sheet.max_row + 1):
                if str(sheet.cell(row=row_index, column=id_column).value or "").strip() != template_id:
                    continue
                return {
                    column: sheet.cell(row=row_index, column=headers[column]).value
                    for column in columns
                }
        finally:
            workbook.close()
        raise AssertionError(f"Template {template_id} not found")

    def test_meta_endpoint_returns_templates_and_decks(self) -> None:
        response = self.client.get("/api/battle/meta")

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertIn("enemyTemplates", payload)
        self.assertIn("decks", payload)
        self.assertIn("playerDecks", payload)
        player_decks_by_id = {item["id"]: item for item in payload["playerDecks"]}
        self.assertIn("human_fighter_lvl1", player_decks_by_id)
        self.assertIn("human_wizzard_lvl1", player_decks_by_id)
        templates_by_id = {item["id"]: item for item in payload["enemyTemplates"]}
        self.assertTrue({"C_GOBLIN", "C_WOLF", "C_HOBGOBLIN", "C_WORG"}.issubset(templates_by_id))
        goblin_template = next(item for item in payload["enemyTemplates"] if item["id"] == "C_GOBLIN")
        self.assertEqual(goblin_template["name"], "Goblin")
        self.assertEqual(goblin_template["imageUrl"], "/images/Changelings/Greenskins/C_GOBLIN.png")
        self.assertEqual(templates_by_id["C_GOBLIN"]["category"], "Changelings")
        self.assertEqual(templates_by_id["C_GOBLIN"]["section"], "Greenskins")
        self.assertTrue(templates_by_id["C_GOBLIN"]["spawnable"])
        self.assertIn("simStats", templates_by_id["C_GOBLIN"])
        self.assertIn("toughness", templates_by_id["C_GOBLIN"]["simStats"])
        self.assertIn("skills", templates_by_id["C_GOBLIN"])
        self.assertEqual(
            templates_by_id["C_GOBLIN"]["simStats"]["initiativeModifier"],
            templates_by_id["C_GOBLIN"]["skills"]["alertness"],
        )
        self.assertIn("simActions", templates_by_id["C_GOBLIN"])
        self.assertIn("effects", templates_by_id["C_GOBLIN"]["simActions"][0])
        self.assertIn("coverageStatus", templates_by_id["C_GOBLIN"]["simActions"][0])
        self.assertEqual(templates_by_id["C_WOLF"]["category"], "Changelings")

    def test_scenario_and_map_template_api_supports_play_runtime(self) -> None:
        map_response = self.client.post(
            "/api/map-templates",
            json={
                "name": "API Map",
                "template": {
                    "tiles": {"0,0": {"tile_type": "floor"}, "1,0": {"tile_type": "floor"}},
                    "walls": {},
                    "rooms": [],
                    "fog_of_war_enabled": False,
                },
            },
        )
        self.assertEqual(map_response.status_code, 200)
        map_id = map_response.json()["template"]["id"]

        loaded_map = self.client.get(f"/api/map-templates/{map_id}")
        self.assertEqual(loaded_map.status_code, 200)
        self.assertIn("0,0", loaded_map.json()["template"]["tiles"])

        updated_map = self.client.put(
            f"/api/map-templates/{map_id}",
            json={
                "name": "API Map Updated",
                "template": {
                    "tiles": {"0,0": {"tile_type": "floor"}},
                    "walls": {},
                    "rooms": [],
                    "fog_of_war_enabled": False,
                },
            },
        )
        self.assertEqual(updated_map.status_code, 200)
        self.assertEqual(updated_map.json()["template"]["name"], "API Map Updated")

        created = self.client.post("/api/scenarios", json={"name": "API Scenario"})
        self.assertEqual(created.status_code, 200)
        scenario_id = created.json()["scenario"]["id"]
        definition = {
            "id": scenario_id,
            "name": "API Scenario",
            "startNodeId": "start",
            "nodes": [
                {
                    "id": "start",
                    "type": "start",
                    "label": "Start",
                    "position": {"x": 0, "y": 0},
                    "defaultPhaseId": "phase_default",
                    "phases": [{"id": "phase_default", "label": "Default", "text": "Start"}],
                },
                {
                    "id": "event",
                    "type": "event",
                    "label": "Event",
                    "position": {"x": 180, "y": 0},
                    "defaultPhaseId": "phase_a",
                    "phases": [
                        {"id": "phase_a", "label": "A", "text": "A"},
                        {"id": "phase_b", "label": "B", "text": "B"},
                    ],
                },
                {
                    "id": "combat",
                    "type": "combat",
                    "label": "Combat",
                    "position": {"x": 360, "y": 0},
                    "defaultPhaseId": "phase_default",
                    "phases": [{"id": "phase_default", "label": "Default", "text": ""}],
                    "combat": {"mapRef": map_id, "enemies": []},
                },
            ],
            "edges": [
                {"id": "edge_start_event", "from": "start", "to": "event", "label": "Investigate"},
                {"id": "edge_event_combat", "from": "event", "to": "combat", "label": "Fight"},
            ],
        }
        saved = self.client.put(f"/api/scenarios/{scenario_id}", json={"definition": definition})
        self.assertEqual(saved.status_code, 200)
        self.assertEqual(saved.json()["scenario"]["edges"][0]["label"], "Investigate")

        sid = self.client.post("/api/battle/sessions").json()["sid"]
        started = self.client.post(
            f"/api/battle/sessions/{sid}/scenario/start-run",
            json={"scenarioId": scenario_id},
        )
        self.assertEqual(started.status_code, 200)
        self.assertEqual(started.json()["scenario"]["runtime"]["currentNodeId"], "start")
        self.assertEqual(started.json()["scenarioRun"]["sourceScenarioId"], scenario_id)
        self.assertEqual(started.json()["scenario"]["definition"]["nodes"][0]["type"], "scene")

        navigated = self.client.post(
            f"/api/battle/sessions/{sid}/scenario/navigate",
            json={"nodeId": "event"},
        )
        self.assertEqual(navigated.status_code, 200)
        self.assertEqual(navigated.json()["scenario"]["runtime"]["currentNodeId"], "event")
        self.assertEqual(navigated.json()["scenario"]["runtime"]["nodeStates"]["event"]["visitCount"], 1)

        phased = self.client.post(
            f"/api/battle/sessions/{sid}/scenario/nodes/event/phase",
            json={"phaseId": "phase_b"},
        )
        self.assertEqual(phased.status_code, 200)
        self.assertEqual(phased.json()["scenario"]["runtime"]["nodeStates"]["event"]["phaseId"], "phase_b")

        updated_definition = {
            **definition,
            "name": "API Scenario Updated",
            "nodes": [
                definition["nodes"][0],
                {**definition["nodes"][1], "label": "Updated Event"},
                definition["nodes"][2],
            ],
        }
        updated_run = self.client.put(
            f"/api/battle/sessions/{sid}/scenario/templates/{scenario_id}",
            json={"definition": updated_definition},
        )
        self.assertEqual(updated_run.status_code, 200)
        self.assertEqual(updated_run.json()["scenarioTemplate"]["name"], "API Scenario Updated")
        self.assertEqual(updated_run.json()["scenario"]["definition"]["nodes"][1]["label"], "Updated Event")
        self.assertEqual(updated_run.json()["scenario"]["runtime"]["currentNodeId"], "event")
        self.assertEqual(updated_run.json()["scenario"]["runtime"]["nodeStates"]["event"]["phaseId"], "phase_b")

        combat = self.client.post(f"/api/battle/sessions/{sid}/scenario/nodes/combat/start-combat")
        self.assertEqual(combat.status_code, 200)
        combat_state = combat.json()["scenario"]["runtime"]["nodeStates"]["combat"]
        self.assertIsNotNone(combat_state["mapInstanceId"])
        self.assertIn("0,0", combat.json()["dungeon"]["tiles"])

    def test_character_builder_catalog_endpoint_returns_classes_and_ancestries(self) -> None:
        response = self.client.get("/api/battle/character-builder/catalog")

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertIn("Martial", payload["energyTypes"])
        class_ids = {entry["id"] for entry in payload["classes"]}
        ancestry_by_id = {entry["id"]: entry for entry in payload["ancestries"]}
        self.assertIn("fighter", class_ids)
        self.assertEqual(ancestry_by_id["halfling"]["card"]["autoDraw"], 1)
        self.assertIn("Disengage without spending an action", ancestry_by_id["halfling"]["card"]["text"])
        art_paths = {entry["imagePath"] for entry in payload["characterArt"]["options"]}
        self.assertIn("Playing_Characters/fighter_human_male.png", art_paths)
        fighter_human = next(entry for entry in payload["characterArt"]["options"] if entry["imagePath"] == "Playing_Characters/fighter_human_male.png")
        self.assertEqual(fighter_human["classId"], "fighter")
        self.assertEqual(fighter_human["ancestryId"], "human")
        self.assertEqual(fighter_human["imageUrl"], "/images/Playing_Characters/fighter_human_male.png")

    def test_character_art_upload_endpoint_stores_custom_images(self) -> None:
        images_dir = Path(self.temp_dir.name) / "images"
        saves_dir = Path(self.temp_dir.name) / "upload_saves"
        context = BattleSessionContext(
            root=PROJECT_ROOT,
            saves_dir=saves_dir,
            images_dir=images_dir,
            map_templates_dir=Path(self.temp_dir.name) / "upload_map_templates",
            scenarios_dir=Path(self.temp_dir.name) / "upload_scenarios",
        )
        app = FastAPI()
        register_battle_api(app, context)
        client = TestClient(app)

        response = client.post(
            "/api/battle/character-builder/art/upload",
            files={"file": ("portrait.png", b"not really png", "image/png")},
        )

        self.assertEqual(response.status_code, 200)
        art = response.json()["art"]
        self.assertEqual(art["source"], "upload")
        self.assertTrue(art["imagePath"].startswith("Playing_Characters/extra/custom/portrait_"))
        self.assertEqual((images_dir / art["imagePath"]).read_bytes(), b"not really png")

        bad_response = client.post(
            "/api/battle/character-builder/art/upload",
            files={"file": ("portrait.txt", b"nope", "text/plain")},
        )
        self.assertEqual(bad_response.status_code, 400)
        self.assertIn("PNG, JPG, JPEG, or WEBP", bad_response.json()["detail"])

    def test_character_profile_crud_and_spawn_endpoint(self) -> None:
        create_response = self.client.post(
            "/api/battle/characters",
            json={
                "name": "Mira",
                "classId": "fighter",
                "ancestryId": "halfling",
                "energyTypes": ["Martial", "Elemental", "Light"],
                "mainArt": "Martial",
                "deckUpgrades": {
                    "Martial": {"success_1": 1, "success_2": 1},
                    "Elemental": {"success_1": 1, "success_2": 1},
                    "Light": {"success_1": 1, "success_2": 1},
                },
                "classImprovementTarget": "success_1",
                "gearPresetId": "melee",
                "art": {
                    "source": "catalog",
                    "imagePath": "Playing_Characters/fighter_human_male.png",
                    "label": "Male",
                },
            },
        )
        self.assertEqual(create_response.status_code, 200)
        character_id = create_response.json()["character"]["id"]
        self.assertEqual(create_response.json()["character"]["art"]["imagePath"], "Playing_Characters/fighter_human_male.png")

        listed = self.client.get("/api/battle/characters").json()["characters"]
        self.assertIn(character_id, {entry["id"] for entry in listed})

        sid = self.client.post("/api/battle/sessions").json()["sid"]
        spawned = self.client.post(
            f"/api/battle/sessions/{sid}/players/from-character",
            json={"characterId": character_id},
        )
        self.assertEqual(spawned.status_code, 200)
        player = next(enemy for enemy in spawned.json()["enemies"] if enemy["template_id"] == "player")
        self.assertEqual(player["name"], "Mira")
        self.assertEqual(player["image_url"], "/images/Playing_Characters/fighter_human_male.png")
        self.assertEqual(player["character_profile"]["className"], "Fighter")
        self.assertEqual(len(player["card_library"]), 20)

        deleted = self.client.delete(f"/api/battle/characters/{character_id}")
        self.assertEqual(deleted.status_code, 200)
        self.assertNotIn(character_id, {entry["id"] for entry in deleted.json()["characters"]})

    def test_character_profile_endpoint_rejects_invalid_energy_without_override(self) -> None:
        response = self.client.post(
            "/api/battle/characters",
            json={
                "name": "Mira",
                "classId": "cleric",
                "ancestryId": "human",
                "energyTypes": ["Martial", "Light", "Shadow"],
                "mainArt": "Light",
                "deckUpgrades": {
                    "Martial": {"success_1": 1, "success_2": 1},
                    "Light": {"success_1": 1, "success_2": 1},
                    "Shadow": {"success_1": 1, "success_2": 1},
                },
                "classImprovementTarget": "success_1",
            },
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn("GM override", response.json()["detail"])

    def test_create_and_load_session(self) -> None:
        create_response = self.client.post("/api/battle/sessions")
        self.assertEqual(create_response.status_code, 200)
        snapshot = create_response.json()
        sid = snapshot["sid"]

        get_response = self.client.get(f"/api/battle/sessions/{sid}")
        self.assertEqual(get_response.status_code, 200)
        self.assertEqual(get_response.json()["sid"], sid)
        self.assertEqual(get_response.json()["round"], 1)
        self.assertEqual(get_response.json()["room"], {"columns": 10, "rows": 7})

    def test_restricted_move_endpoint_tracks_pool_and_position_repositions_freely(self) -> None:
        sid = self.client.post("/api/battle/sessions").json()["sid"]
        added = self.client.post(f"/api/battle/sessions/{sid}/enemies", json={"templateId": "C_GOBLIN"}).json()
        entity_id = added["selectedId"]
        self.client.post(f"/api/battle/sessions/{sid}/entities/{entity_id}/position", json={"x": 0, "y": 0})
        self.client.post(f"/api/battle/sessions/{sid}/turn/next")
        self.client.post(f"/api/battle/sessions/{sid}/round/start")  # single unit wraps

        moved = self.client.post(f"/api/battle/sessions/{sid}/entities/{entity_id}/move", json={"x": 2, "y": 0})
        self.assertEqual(moved.status_code, 200)
        self.assertEqual(moved.json()["movementState"]["movementUsed"], 2)
        self.assertFalse(moved.json()["movementState"]["dashUsed"])

        repositioned = self.client.post(f"/api/battle/sessions/{sid}/entities/{entity_id}/position", json={"x": 5, "y": 0})
        self.assertEqual(repositioned.status_code, 200)
        self.assertEqual(repositioned.json()["movementState"]["movementUsed"], 2)
        self.client.post(
            f"/api/battle/sessions/{sid}/dungeon/tiles",
            json={"tileType": "floor", "cells": [[10, 0]]},
        )

        dash_required = self.client.post(f"/api/battle/sessions/{sid}/entities/{entity_id}/move", json={"x": 10, "y": 0})
        self.assertEqual(dash_required.status_code, 400)
        self.assertIn("Dash", dash_required.json()["detail"])

        dashed = self.client.post(
            f"/api/battle/sessions/{sid}/entities/{entity_id}/move",
            json={"x": 10, "y": 0, "dash": True},
        )
        self.assertEqual(dashed.status_code, 200)
        self.assertEqual(dashed.json()["movementState"]["movementUsed"], 7)
        self.assertTrue(dashed.json()["movementState"]["dashUsed"])

    def test_position_endpoint_accepts_negative_sparse_floor(self) -> None:
        sid = self.client.post("/api/battle/sessions").json()["sid"]
        added = self.client.post(f"/api/battle/sessions/{sid}/enemies", json={"templateId": "C_GOBLIN"}).json()
        entity_id = added["selectedId"]
        tile_response = self.client.post(
            f"/api/battle/sessions/{sid}/dungeon/tiles",
            json={"tileType": "floor", "cells": [[-1, 0]]},
        )
        self.assertEqual(tile_response.status_code, 200)

        positioned = self.client.post(
            f"/api/battle/sessions/{sid}/entities/{entity_id}/position",
            json={"x": -1, "y": 0},
        )
        self.assertEqual(positioned.status_code, 200)
        moved = next(enemy for enemy in positioned.json()["enemies"] if enemy["instance_id"] == entity_id)
        self.assertEqual((moved["grid_x"], moved["grid_y"]), (-1, 0))

        missing_tile = self.client.post(
            f"/api/battle/sessions/{sid}/entities/{entity_id}/position",
            json={"x": -2, "y": 0},
        )
        self.assertEqual(missing_tile.status_code, 400)
        self.assertIn("not walkable", missing_tile.json()["detail"])

    def test_batch_position_endpoint_moves_group_atomically_and_undoes_once(self) -> None:
        sid = self.client.post("/api/battle/sessions").json()["sid"]
        first_id = self.client.post(f"/api/battle/sessions/{sid}/enemies", json={"templateId": "C_GOBLIN"}).json()["selectedId"]
        second_id = self.client.post(f"/api/battle/sessions/{sid}/enemies", json={"templateId": "C_WOLF"}).json()["selectedId"]
        blocker_id = self.client.post(f"/api/battle/sessions/{sid}/enemies", json={"templateId": "C_HOBGOBLIN"}).json()["selectedId"]
        self.client.post(f"/api/battle/sessions/{sid}/entities/{first_id}/position", json={"x": 0, "y": 0})
        self.client.post(f"/api/battle/sessions/{sid}/entities/{second_id}/position", json={"x": 1, "y": 0})
        self.client.post(f"/api/battle/sessions/{sid}/entities/{blocker_id}/position", json={"x": 3, "y": 0})
        before = self.client.get(f"/api/battle/sessions/{sid}").json()

        blocked = self.client.post(
            f"/api/battle/sessions/{sid}/entities/positions",
            json={
                "placements": [
                    {"instanceId": first_id, "x": 1, "y": 0},
                    {"instanceId": second_id, "x": 3, "y": 0},
                ],
            },
        )
        self.assertEqual(blocked.status_code, 400)
        unchanged = self.client.get(f"/api/battle/sessions/{sid}").json()
        positions = {enemy["instance_id"]: (enemy["grid_x"], enemy["grid_y"]) for enemy in unchanged["enemies"]}
        self.assertEqual(positions[first_id], (0, 0))
        self.assertEqual(positions[second_id], (1, 0))
        self.assertEqual(unchanged["undoDepth"], before["undoDepth"])

        moved = self.client.post(
            f"/api/battle/sessions/{sid}/entities/positions",
            json={
                "placements": [
                    {"instanceId": first_id, "x": 1, "y": 0},
                    {"instanceId": second_id, "x": 0, "y": 0},
                ],
            },
        )
        self.assertEqual(moved.status_code, 200)
        moved_payload = moved.json()
        self.assertEqual(moved_payload["undoDepth"], before["undoDepth"] + 1)
        moved_positions = {enemy["instance_id"]: (enemy["grid_x"], enemy["grid_y"]) for enemy in moved_payload["enemies"]}
        self.assertEqual(moved_positions[first_id], (1, 0))
        self.assertEqual(moved_positions[second_id], (0, 0))

        undone = self.client.post(f"/api/battle/sessions/{sid}/undo").json()
        undone_positions = {enemy["instance_id"]: (enemy["grid_x"], enemy["grid_y"]) for enemy in undone["enemies"]}
        self.assertEqual(undone_positions[first_id], (0, 0))
        self.assertEqual(undone_positions[second_id], (1, 0))

    def test_party_walk_endpoint_moves_player_party_and_returns_payload(self) -> None:
        sid = self.client.post("/api/battle/sessions").json()["sid"]
        first_id = self.client.post(f"/api/battle/sessions/{sid}/players", json={"name": "Leader"}).json()["selectedId"]
        second_id = self.client.post(f"/api/battle/sessions/{sid}/players", json={"name": "Follower"}).json()["selectedId"]
        self.client.post(f"/api/battle/sessions/{sid}/entities/{first_id}/position", json={"x": 0, "y": 1})
        self.client.post(f"/api/battle/sessions/{sid}/entities/{second_id}/position", json={"x": 0, "y": 2})

        response = self.client.post(
            f"/api/battle/sessions/{sid}/action/party-walk",
            json={"leaderId": first_id, "x": 4, "y": 1},
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["partyWalk"]["leaderId"], first_id)
        self.assertEqual(payload["partyWalk"]["destination"], {"x": 4, "y": 1})
        self.assertEqual(payload["partyWalk"]["actualDestination"], {"x": 4, "y": 1})
        self.assertEqual(set(payload["partyWalk"]["movedEntityIds"]), {first_id, second_id})
        positions = {enemy["instance_id"]: (enemy["grid_x"], enemy["grid_y"]) for enemy in payload["enemies"]}
        self.assertEqual(positions[first_id], (4, 1))
        self.assertEqual(positions[second_id], (3, 1))

    def test_party_walk_endpoint_rejects_enemy_leader(self) -> None:
        sid = self.client.post("/api/battle/sessions").json()["sid"]
        enemy_id = self.client.post(f"/api/battle/sessions/{sid}/enemies", json={"templateId": "C_GOBLIN"}).json()["selectedId"]

        response = self.client.post(
            f"/api/battle/sessions/{sid}/action/party-walk",
            json={"leaderId": enemy_id, "x": 0, "y": 0},
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn("player character", response.json()["detail"])

    def test_walk_endpoint_moves_single_unit_and_returns_payload(self) -> None:
        sid = self.client.post("/api/battle/sessions").json()["sid"]
        first_id = self.client.post(f"/api/battle/sessions/{sid}/players", json={"name": "Walker"}).json()["selectedId"]
        second_id = self.client.post(f"/api/battle/sessions/{sid}/players", json={"name": "Follower"}).json()["selectedId"]
        self.client.post(f"/api/battle/sessions/{sid}/entities/{first_id}/position", json={"x": 0, "y": 1})
        self.client.post(f"/api/battle/sessions/{sid}/entities/{second_id}/position", json={"x": 0, "y": 2})

        response = self.client.post(
            f"/api/battle/sessions/{sid}/entities/{first_id}/walk",
            json={"x": 4, "y": 1},
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["walk"]["entityId"], first_id)
        self.assertEqual(payload["walk"]["destination"], {"x": 4, "y": 1})
        self.assertEqual(payload["walk"]["actualDestination"], {"x": 4, "y": 1})
        positions = {enemy["instance_id"]: (enemy["grid_x"], enemy["grid_y"]) for enemy in payload["enemies"]}
        self.assertEqual(positions[first_id], (4, 1))
        self.assertEqual(positions[second_id], (0, 2))

    def test_copy_endpoint_creates_fresh_enemy_and_is_undoable(self) -> None:
        sid = self.client.post("/api/battle/sessions").json()["sid"]
        added = self.client.post(f"/api/battle/sessions/{sid}/enemies", json={"templateId": "C_GOBLIN"}).json()
        source_id = added["selectedId"]
        before = self.client.get(f"/api/battle/sessions/{sid}").json()
        source = next(enemy for enemy in before["enemies"] if enemy["instance_id"] == source_id)

        copied = self.client.post(f"/api/battle/sessions/{sid}/entities/{source_id}/copy")

        self.assertEqual(copied.status_code, 200)
        payload = copied.json()
        self.assertEqual(payload["undoDepth"], before["undoDepth"] + 1)
        self.assertEqual(len(payload["enemies"]), len(before["enemies"]) + 1)
        copy = next(enemy for enemy in payload["enemies"] if enemy["instance_id"] == payload["selectedId"])
        self.assertNotEqual(copy["instance_id"], source_id)
        self.assertEqual(copy["template_id"], source["template_id"])
        self.assertEqual((copy["grid_x"], copy["grid_y"]), (source["grid_x"] + 1, source["grid_y"]))
        self.assertEqual(payload["order"], [source_id, copy["instance_id"]])

        undone = self.client.post(f"/api/battle/sessions/{sid}/undo").json()
        self.assertEqual([enemy["instance_id"] for enemy in undone["enemies"]], [source_id])

    def test_entity_loot_inspect_endpoint_targets_path_entity(self) -> None:
        sid = self.client.post("/api/battle/sessions").json()["sid"]
        goblin_id = self.client.post(f"/api/battle/sessions/{sid}/enemies", json={"templateId": "C_GOBLIN"}).json()["selectedId"]
        wolf_id = self.client.post(f"/api/battle/sessions/{sid}/enemies", json={"templateId": "C_WOLF"}).json()["selectedId"]
        session = self.context.load_session(sid)
        session.state.enemies[goblin_id].toughness_current = 0
        session.selected_id = wolf_id
        session.autosave()

        response = self.client.post(f"/api/battle/sessions/{sid}/entities/{goblin_id}/loot/inspect")

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["selectedId"], goblin_id)
        goblin = next(enemy for enemy in payload["enemies"] if enemy["instance_id"] == goblin_id)
        wolf = next(enemy for enemy in payload["enemies"] if enemy["instance_id"] == wolf_id)
        self.assertTrue(goblin["loot_rolled"])
        self.assertEqual(goblin["loot_state"], "inspected")
        self.assertFalse(wolf["loot_rolled"])

    def test_take_loot_endpoint_adds_inventory_and_rejects_double_take(self) -> None:
        sid = self.client.post("/api/battle/sessions").json()["sid"]
        player_id = self.client.post(f"/api/battle/sessions/{sid}/players", json={"name": "Mira"}).json()["selectedId"]
        enemy_id = self.client.post(f"/api/battle/sessions/{sid}/enemies", json={"templateId": "C_GOBLIN"}).json()["selectedId"]
        session = self.context.load_session(sid)
        player = session.state.enemies[player_id]
        enemy = session.state.enemies[enemy_id]
        player.grid_x, player.grid_y = 5, 3
        enemy.grid_x, enemy.grid_y = 4, 3
        enemy.toughness_current = 0
        enemy.loot_rolled = True
        enemy.rolled_loot = {"currency": {"cp": 4}, "resources": {}, "other": ["note"]}
        session.autosave()

        response = self.client.post(
            f"/api/battle/sessions/{sid}/entities/{enemy_id}/loot/take",
            json={"playerId": player_id},
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["selectedId"], player_id)
        player_payload = next(enemy for enemy in payload["enemies"] if enemy["instance_id"] == player_id)
        enemy_payload = next(enemy for enemy in payload["enemies"] if enemy["instance_id"] == enemy_id)
        self.assertEqual(player_payload["inventory"]["currency"]["cp"], 4)
        self.assertEqual(player_payload["inventory"]["other"], ["note"])
        self.assertEqual(enemy_payload["loot_state"], "taken")
        self.assertEqual(enemy_payload["loot_taken_by_name"], "Mira")

        rejected = self.client.post(
            f"/api/battle/sessions/{sid}/entities/{enemy_id}/loot/take",
            json={"playerId": player_id},
        )
        self.assertEqual(rejected.status_code, 400)
        self.assertIn("already been taken", rejected.json()["detail"])

    def test_take_loot_endpoint_charges_action_in_combat(self) -> None:
        sid = self.client.post("/api/battle/sessions").json()["sid"]
        player_id = self.client.post(f"/api/battle/sessions/{sid}/players", json={"name": "Mira"}).json()["selectedId"]
        enemy_id = self.client.post(f"/api/battle/sessions/{sid}/enemies", json={"templateId": "C_GOBLIN"}).json()["selectedId"]
        session = self.context.load_session(sid)
        player = session.state.enemies[player_id]
        enemy = session.state.enemies[enemy_id]
        player.grid_x, player.grid_y = 5, 3
        enemy.grid_x, enemy.grid_y = 4, 3
        enemy.toughness_current = 0
        enemy.loot_rolled = True
        enemy.rolled_loot = {"currency": {"cp": 2}, "resources": {}, "other": []}
        session.encounter_started = True
        session.active_turn_id = player_id
        session.autosave()

        response = self.client.post(
            f"/api/battle/sessions/{sid}/entities/{enemy_id}/loot/take",
            json={"playerId": player_id},
        )

        self.assertEqual(response.status_code, 200)
        player_payload = next(enemy for enemy in response.json()["enemies"] if enemy["instance_id"] == player_id)
        self.assertEqual(player_payload["actions_used"], 1)

    def test_inspect_all_loot_endpoint_skips_hidden_fog_rooms(self) -> None:
        sid = self.client.post("/api/battle/sessions").json()["sid"]
        session = self.context.load_session(sid)
        session.edit_dungeon_tiles("void", [[x, y] for x in range(10) for y in range(7)])
        session.edit_dungeon_tiles("floor", [[0, 0], [1, 0], [4, 0]])
        session.analyze_dungeon()
        visible_room = next(room.room_id for room in session.dungeon.rooms if any(tuple(cell) == (0, 0) for cell in room.cells))
        hidden_room = next(room.room_id for room in session.dungeon.rooms if any(tuple(cell) == (4, 0) for cell in room.cells))
        session.add_player(name="Mira")
        player = session.state.enemies[session.selected_id]
        player.grid_x, player.grid_y = 0, 0
        player.room_id = visible_room
        session.add_enemy_from_template("C_GOBLIN")
        visible_enemy_id = session.selected_id
        visible_enemy = session.state.enemies[visible_enemy_id]
        visible_enemy.grid_x, visible_enemy.grid_y = 1, 0
        visible_enemy.room_id = visible_room
        visible_enemy.toughness_current = 0
        session.add_enemy_from_template("C_GOBLIN")
        hidden_enemy_id = session.selected_id
        hidden_enemy = session.state.enemies[hidden_enemy_id]
        hidden_enemy.grid_x, hidden_enemy.grid_y = 4, 0
        hidden_enemy.room_id = hidden_room
        hidden_enemy.toughness_current = 0
        session.dungeon.fog_of_war_enabled = True
        session.dungeon.revealed_room_ids = [visible_room]
        session.autosave()

        response = self.client.post(f"/api/battle/sessions/{sid}/loot/inspect-all")

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        visible_payload = next(enemy for enemy in payload["enemies"] if enemy["instance_id"] == visible_enemy_id)
        hidden_payload = next(enemy for enemy in payload["enemies"] if enemy["instance_id"] == hidden_enemy_id)
        self.assertTrue(visible_payload["loot_rolled"])
        self.assertFalse(hidden_payload["loot_rolled"])

    def test_suspect_interaction_resolves_willpower_and_undoes_as_one_check(self) -> None:
        sid = self.client.post("/api/battle/sessions").json()["sid"]
        session = self.context.load_session(sid)
        session.edit_dungeon_tiles("void", [[x, y] for x in range(10) for y in range(7)])
        session.edit_dungeon_tiles("floor", [[0, 0], [0, 1], [1, 0]])
        session.analyze_dungeon()
        session.add_player()
        player = session.state.enemies[session.selected_id]
        player.grid_x, player.grid_y = 0, 0
        player.room_id = session.dungeon.rooms[0].room_id
        player.deck_state.hand = []
        player.deck_state.discard_pile = []
        player.deck_state.draw_pile = ["hf_martial_1_fail", "hf_void_fate", "hf_void_fail"]
        session.dungeon.secret_suspects = [{
            "room_id": player.room_id,
            "edge_key": "0,0,e",
            "kind": "false",
            "exhausted": False,
            "false_dc": 1,
        }]
        session.autosave()

        started = self.client.post(
            f"/api/battle/sessions/{sid}/dungeon/suspects/interact",
            json={"edgeKey": "0,0,e"},
        ).json()
        self.assertEqual(started["pendingSearch"]["kind"], "suspect")
        self.assertTrue(started["pendingSearch"]["hasFate"])
        self.assertEqual(started["undoDepth"], 1)

        resolved = self.client.post(
            f"/api/battle/sessions/{sid}/dungeon/suspects/resolve",
            json={"useWillpower": True},
        ).json()
        self.assertIsNone(resolved["pendingSearch"])
        self.assertEqual(resolved["suspectInteraction"]["outcome"], "cleared")
        self.assertEqual(resolved["undoDepth"], 1)
        self.assertEqual(resolved["dungeon"]["secretSuspects"], [])

        undone = self.client.post(f"/api/battle/sessions/{sid}/undo").json()
        self.assertIsNone(undone["pendingSearch"])
        self.assertEqual(undone["undoDepth"], 0)
        self.assertEqual(len(undone["dungeon"]["secretSuspects"]), 1)

    def test_room_search_resolve_defaults_party_walk_false(self) -> None:
        sid = self.client.post("/api/battle/sessions").json()["sid"]
        session = self.context.load_session(sid)
        session.edit_dungeon_tiles("void", [[x, y] for x in range(10) for y in range(7)])
        session.edit_dungeon_tiles("floor", [[0, 0], [0, 1], [1, 0]])
        session.edit_dungeon_walls("secret_door", [{"x": 0, "y": 0, "side": "e"}])
        session.analyze_dungeon()
        session.dungeon.walls["0,0,e"].secret_dc = 1
        session.add_player()
        player = session.state.enemies[session.selected_id]
        player.grid_x, player.grid_y = 0, 1
        player.room_id = session._room_id_for_position(0, 1)
        player.deck_state.hand = []
        player.deck_state.discard_pile = []
        player.deck_state.draw_pile = ["hf_martial_success_2", "hf_void_fail", "hf_void_fail"]
        session.autosave()

        self.client.post(f"/api/battle/sessions/{sid}/dungeon/search/start")
        response = self.client.post(
            f"/api/battle/sessions/{sid}/dungeon/search/resolve",
            json={"useWillpower": False},
        )

        self.assertEqual(response.status_code, 200)
        resolved = response.json()["searchResolved"]
        self.assertFalse(resolved["partyWalk"])
        self.assertEqual(resolved["edgeKey"], "0,0,e")
        self.assertEqual(resolved["edgeKeys"], ["0,0,e"])
        self.assertIn(player.instance_id, resolved["movedEntityIds"])

    def test_room_search_resolve_accepts_party_walk_true(self) -> None:
        sid = self.client.post("/api/battle/sessions").json()["sid"]
        session = self.context.load_session(sid)
        session.edit_dungeon_tiles("void", [[x, y] for x in range(10) for y in range(7)])
        session.edit_dungeon_tiles("floor", [[0, 0], [0, 1], [1, 0]])
        session.edit_dungeon_walls("secret_door", [{"x": 0, "y": 0, "side": "e"}])
        session.analyze_dungeon()
        session.dungeon.walls["0,0,e"].secret_dc = 1
        ids = []
        for position in [(0, 1), (0, 0)]:
            session.add_player()
            entity = session.state.enemies[session.selected_id]
            session._set_position(entity, position[0], position[1])
            ids.append(session.selected_id)
        searcher = session.state.enemies[ids[0]]
        searcher.deck_state.hand = []
        searcher.deck_state.discard_pile = []
        searcher.deck_state.draw_pile = ["hf_martial_success_2", "hf_void_fail", "hf_void_fail"]
        session.selected_id = ids[0]
        session.autosave()

        self.client.post(f"/api/battle/sessions/{sid}/dungeon/search/start")
        response = self.client.post(
            f"/api/battle/sessions/{sid}/dungeon/search/resolve",
            json={"useWillpower": False, "partyWalk": True},
        )

        self.assertEqual(response.status_code, 200)
        resolved = response.json()["searchResolved"]
        self.assertTrue(resolved["partyWalk"])
        self.assertEqual(resolved["edgeKeys"], ["0,0,e"])
        self.assertEqual(set(resolved["movedEntityIds"]), set(ids))

    def test_start_encounter_endpoint_activates_highest_initiative_unit(self) -> None:
        sid = self.client.post("/api/battle/sessions").json()["sid"]
        first_enemy = self.client.post(f"/api/battle/sessions/{sid}/enemies", json={"templateId": "C_GOBLIN"}).json()
        first_id = first_enemy["selectedId"]
        self.client.post(f"/api/battle/sessions/{sid}/enemies", json={"templateId": "C_WOLF"})

        response = self.client.post(f"/api/battle/sessions/{sid}/encounter/start")

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        active_enemy = next(enemy for enemy in payload["enemies"] if enemy["instance_id"] == first_id)
        self.assertEqual(payload["activeTurnId"], first_id)
        self.assertEqual(payload["selectedId"], first_id)
        self.assertFalse(payload["turnInProgress"])
        self.assertEqual(payload["movementState"]["entityId"], first_id)
        self.assertEqual(active_enemy["current_draw_text"], [])

    def test_enemy_flow_and_manual_save_endpoints(self) -> None:
        snapshot = self.client.post("/api/battle/sessions").json()
        sid = snapshot["sid"]

        add_enemy = self.client.post(f"/api/battle/sessions/{sid}/enemies", json={"templateId": "C_GOBLIN"})
        self.assertEqual(add_enemy.status_code, 200)
        entity_id = add_enemy.json()["selectedId"]

        draw_response = self.client.post(f"/api/battle/sessions/{sid}/turn/draw")
        self.assertEqual(draw_response.status_code, 200)
        self.assertIsNotNone(draw_response.json()["activeTurnId"])

        save_response = self.client.post(f"/api/battle/sessions/{sid}/saves", json={"name": "api-save"})
        self.assertEqual(save_response.status_code, 200)
        save_payload = save_response.json()
        self.assertIn("Session save created", save_payload["combatLog"][0])
        self.assertEqual(save_payload["save"]["name"], "api-save")
        self.assertEqual(save_payload["activeSave"]["filename"], save_payload["save"]["filename"])

        saves_response = self.client.get(f"/api/battle/sessions/{sid}/saves")
        self.assertEqual(saves_response.status_code, 200)
        saves = saves_response.json()["saves"]
        self.assertEqual(len(saves), 1)
        self.assertTrue(saves[0]["active"])
        self.assertEqual(saves[0]["name"], "api-save")

        self.client.post(f"/api/battle/sessions/{sid}/enemies", json={"templateId": "C_WOLF"})
        overwrite_response = self.client.put(f"/api/battle/sessions/{sid}/saves/{saves[0]['filename']}")
        self.assertEqual(overwrite_response.status_code, 200)
        self.assertEqual(overwrite_response.json()["save"]["filename"], saves[0]["filename"])
        self.assertEqual(overwrite_response.json()["activeSave"]["filename"], saves[0]["filename"])

        delete_response = self.client.delete(f"/api/battle/sessions/{sid}/entities/{entity_id}")
        self.assertEqual(delete_response.status_code, 200)
        self.assertEqual(len(delete_response.json()["order"]), 1)

        load_response = self.client.post(
            f"/api/battle/sessions/{sid}/load",
            json={"filename": saves[0]["filename"]},
        )
        self.assertEqual(load_response.status_code, 200)
        self.assertEqual(len(load_response.json()["order"]), 2)
        self.assertEqual(load_response.json()["activeSave"]["filename"], saves[0]["filename"])

    def test_manual_save_delete_endpoint(self) -> None:
        snapshot = self.client.post("/api/battle/sessions").json()
        sid = snapshot["sid"]

        save_response = self.client.post(f"/api/battle/sessions/{sid}/saves", json={"name": "old-load"})
        self.assertEqual(save_response.status_code, 200)
        saves = self.client.get(f"/api/battle/sessions/{sid}/saves").json()["saves"]
        self.assertEqual(len(saves), 1)

        delete_response = self.client.delete(f"/api/battle/sessions/{sid}/saves/{saves[0]['filename']}")

        self.assertEqual(delete_response.status_code, 200)
        self.assertEqual(delete_response.json()["saves"], [])
        self.assertIsNone(delete_response.json()["activeSave"])
        self.assertEqual(self.client.get(f"/api/battle/sessions/{sid}/saves").json()["saves"], [])

    def test_new_taxonomy_templates_can_be_added(self) -> None:
        sid = self.client.post("/api/battle/sessions").json()["sid"]

        for template_id in ("C_HOBGOBLIN", "C_WORG"):
            response = self.client.post(f"/api/battle/sessions/{sid}/enemies", json={"templateId": template_id})
            self.assertEqual(response.status_code, 200)
            selected = next(enemy for enemy in response.json()["enemies"] if enemy["instance_id"] == response.json()["selectedId"])
            self.assertEqual(selected["template_id"], template_id)

    def test_next_clears_current_draw_until_a_new_draw_happens(self) -> None:
        sid = self.client.post("/api/battle/sessions").json()["sid"]
        add_enemy = self.client.post(f"/api/battle/sessions/{sid}/enemies", json={"templateId": "C_GOBLIN"}).json()
        entity_id = add_enemy["selectedId"]

        draw_response = self.client.post(f"/api/battle/sessions/{sid}/turn/draw")
        self.assertEqual(draw_response.status_code, 200)
        drawn_enemy = next(enemy for enemy in draw_response.json()["enemies"] if enemy["instance_id"] == entity_id)
        self.assertGreaterEqual(len(drawn_enemy["current_draw_text"]), 1)

        end_response = self.client.post(f"/api/battle/sessions/{sid}/turn/end")
        self.assertEqual(end_response.status_code, 200)
        ended_enemy = next(enemy for enemy in end_response.json()["enemies"] if enemy["instance_id"] == entity_id)
        self.assertGreaterEqual(len(ended_enemy["current_draw_text"]), 1)

        self.client.post(f"/api/battle/sessions/{sid}/turn/next")  # single unit wraps
        start_response = self.client.post(f"/api/battle/sessions/{sid}/round/start")
        self.assertEqual(start_response.status_code, 200)
        next_payload = start_response.json()
        next_enemy = next(enemy for enemy in next_payload["enemies"] if enemy["instance_id"] == entity_id)

        self.assertEqual(next_payload["activeTurnId"], entity_id)
        self.assertFalse(next_payload["turnInProgress"])
        self.assertEqual(next_enemy["current_draw_text"], [])

    def test_next_to_other_unit_keeps_previous_units_draw_visible(self) -> None:
        sid = self.client.post("/api/battle/sessions").json()["sid"]
        first_enemy = self.client.post(f"/api/battle/sessions/{sid}/enemies", json={"templateId": "C_GOBLIN"}).json()
        first_id = first_enemy["selectedId"]
        second_enemy = self.client.post(f"/api/battle/sessions/{sid}/enemies", json={"templateId": "C_WOLF"}).json()
        second_id = second_enemy["selectedId"]

        self.client.post(f"/api/battle/sessions/{sid}/select", json={"instanceId": first_id})
        draw_response = self.client.post(f"/api/battle/sessions/{sid}/turn/draw")
        self.assertEqual(draw_response.status_code, 200)
        self.client.post(f"/api/battle/sessions/{sid}/turn/end")

        next_response = self.client.post(f"/api/battle/sessions/{sid}/turn/next")
        self.assertEqual(next_response.status_code, 200)
        next_payload = next_response.json()
        self.assertEqual(next_payload["activeTurnId"], second_id)

        self.client.post(f"/api/battle/sessions/{sid}/select", json={"instanceId": first_id})
        selected_first = self.client.get(f"/api/battle/sessions/{sid}").json()
        first_enemy_payload = next(enemy for enemy in selected_first["enemies"] if enemy["instance_id"] == first_id)
        self.assertGreaterEqual(len(first_enemy_payload["current_draw_text"]), 1)

    def test_attack_and_heal_endpoints_return_updated_snapshot(self) -> None:
        sid = self.client.post("/api/battle/sessions").json()["sid"]
        enemy_snapshot = self.client.post(f"/api/battle/sessions/{sid}/enemies", json={"templateId": "C_WOLF"}).json()

        before = next(enemy for enemy in enemy_snapshot["enemies"] if enemy["instance_id"] == enemy_snapshot["selectedId"])

        attack_response = self.client.post(
            f"/api/battle/sessions/{sid}/attack",
            json={"damage": 2, "burn": True, "poison": False, "slow": False, "paralyze": False, "modifiers": []},
        )
        self.assertEqual(attack_response.status_code, 200)
        attacked = next(
            enemy for enemy in attack_response.json()["enemies"] if enemy["instance_id"] == attack_response.json()["selectedId"]
        )
        self.assertLessEqual(attacked["toughness_current"], before["toughness_current"])
        self.assertIn("burn", attacked["statuses"])

        heal_response = self.client.post(
            f"/api/battle/sessions/{sid}/heal",
            json={"toughness": 1, "armor": 0, "magicArmor": 0, "C_HOBGOBLIN": 0},
        )
        self.assertEqual(heal_response.status_code, 200)
        healed = next(enemy for enemy in heal_response.json()["enemies"] if enemy["instance_id"] == enemy_snapshot["selectedId"])
        self.assertGreaterEqual(healed["toughness_current"], attacked["toughness_current"])

    def test_attack_endpoint_reports_player_wounds(self) -> None:
        sid = self.client.post("/api/battle/sessions").json()["sid"]
        player_snapshot = self.client.post(
            f"/api/battle/sessions/{sid}/players",
            json={"name": "Mira", "toughness": 5, "armor": 0, "magicArmor": 0, "power": 0, "movement": 6},
        ).json()

        attack_response = self.client.post(
            f"/api/battle/sessions/{sid}/attack",
            json={"damage": 11, "burn": False, "poison": False, "slow": False, "paralyze": False, "modifiers": []},
        )

        self.assertEqual(attack_response.status_code, 200)
        payload = attack_response.json()
        player = next(enemy for enemy in payload["enemies"] if enemy["instance_id"] == player_snapshot["selectedId"])
        self.assertEqual(player["toughness_current"], 4)
        self.assertFalse(player["is_down"])
        self.assertFalse(player["is_ko"])
        self.assertEqual(player["wound_counts"], {"hand": 2, "discard": 0, "draw_pile": 0, "total": 2})
        self.assertEqual(payload["woundEvents"][0]["name"], "Mira")
        self.assertEqual(payload["woundEvents"][0]["wounds"], 2)
        self.assertEqual(payload["woundEvents"][0]["toughnessAfter"], 4)

    def test_player_wound_endpoints_discard_remove_and_require_deck_confirmation(self) -> None:
        sid = self.client.post("/api/battle/sessions").json()["sid"]
        player_snapshot = self.client.post(
            f"/api/battle/sessions/{sid}/players",
            json={"name": "Mira", "power": 1},
        ).json()
        player_id = player_snapshot["selectedId"]
        session = self.context.load_session(sid)
        player = session.state.enemies[player_id]
        player.deck_state.hand = [WOUND_CARD_ID]
        player.deck_state.discard_pile = [WOUND_CARD_ID]
        player.deck_state.draw_pile = [WOUND_CARD_ID, "hf_martial_1_success"]
        session.autosave()

        discarded = self.client.post(f"/api/battle/sessions/{sid}/entities/{player_id}/wounds/discard")
        self.assertEqual(discarded.status_code, 200)
        player_payload = next(enemy for enemy in discarded.json()["enemies"] if enemy["instance_id"] == player_id)
        self.assertEqual(player_payload["wound_counts"], {"hand": 0, "discard": 2, "draw_pile": 1, "total": 3})

        removed_discard = self.client.post(
            f"/api/battle/sessions/{sid}/entities/{player_id}/wounds/remove",
            json={"confirmDeck": False},
        )
        self.assertEqual(removed_discard.status_code, 200)
        player_payload = next(enemy for enemy in removed_discard.json()["enemies"] if enemy["instance_id"] == player_id)
        self.assertEqual(player_payload["wound_counts"], {"hand": 0, "discard": 1, "draw_pile": 1, "total": 2})

        self.client.post(f"/api/battle/sessions/{sid}/entities/{player_id}/wounds/remove", json={"confirmDeck": False})
        unconfirmed = self.client.post(
            f"/api/battle/sessions/{sid}/entities/{player_id}/wounds/remove",
            json={"confirmDeck": False},
        )
        self.assertEqual(unconfirmed.status_code, 400)
        self.assertIn("requires confirmation", unconfirmed.json()["detail"])

        confirmed = self.client.post(
            f"/api/battle/sessions/{sid}/entities/{player_id}/wounds/remove",
            json={"confirmDeck": True},
        )
        self.assertEqual(confirmed.status_code, 200)
        player_payload = next(enemy for enemy in confirmed.json()["enemies"] if enemy["instance_id"] == player_id)
        self.assertEqual(player_payload["wound_counts"], {"hand": 0, "discard": 0, "draw_pile": 0, "total": 0})

    def test_physical_player_cards_api_tracks_total_wounds_and_rejects_digital_draw(self) -> None:
        sid = self.client.post("/api/battle/sessions").json()["sid"]
        player_snapshot = self.client.post(
            f"/api/battle/sessions/{sid}/players",
            json={"name": "Mira", "toughness": 5, "armor": 0, "magicArmor": 0, "power": 1, "physicalCards": True},
        ).json()
        player_id = player_snapshot["selectedId"]
        player_payload = next(enemy for enemy in player_snapshot["enemies"] if enemy["instance_id"] == player_id)
        self.assertTrue(player_payload["physical_cards"])

        attack_response = self.client.post(
            f"/api/battle/sessions/{sid}/attack",
            json={"damage": 11, "burn": False, "poison": False, "slow": False, "paralyze": False, "modifiers": []},
        )

        self.assertEqual(attack_response.status_code, 200)
        player_payload = next(enemy for enemy in attack_response.json()["enemies"] if enemy["instance_id"] == player_id)
        self.assertEqual(player_payload["wound_counts"], {"hand": 0, "discard": 0, "draw_pile": 0, "total": 2})

        draw_response = self.client.post(f"/api/battle/sessions/{sid}/turn/draw")
        self.assertEqual(draw_response.status_code, 400)
        self.assertIn("outside the app", draw_response.json()["detail"])

        adjusted = self.client.post(
            f"/api/battle/sessions/{sid}/entities/{player_id}/wounds/adjust",
            json={"delta": -1},
        )
        self.assertEqual(adjusted.status_code, 200)
        player_payload = next(enemy for enemy in adjusted.json()["enemies"] if enemy["instance_id"] == player_id)
        self.assertEqual(player_payload["wound_counts"]["total"], 1)

    def test_player_card_mode_endpoint_converts_digital_wounds(self) -> None:
        sid = self.client.post("/api/battle/sessions").json()["sid"]
        player_snapshot = self.client.post(
            f"/api/battle/sessions/{sid}/players",
            json={"name": "Mira", "power": 1},
        ).json()
        player_id = player_snapshot["selectedId"]
        session = self.context.load_session(sid)
        player = session.state.enemies[player_id]
        player.deck_state.hand = [WOUND_CARD_ID]
        player.deck_state.discard_pile = [WOUND_CARD_ID]
        player.deck_state.draw_pile = [WOUND_CARD_ID, "hf_martial_1_success"]
        session.autosave()

        converted = self.client.post(
            f"/api/battle/sessions/{sid}/entities/{player_id}/player-card-mode",
            json={"physicalCards": True},
        )

        self.assertEqual(converted.status_code, 200)
        player_payload = next(enemy for enemy in converted.json()["enemies"] if enemy["instance_id"] == player_id)
        self.assertTrue(player_payload["physical_cards"])
        self.assertEqual(player_payload["wound_counts"], {"hand": 0, "discard": 0, "draw_pile": 0, "total": 3})

        disabled = self.client.post(
            f"/api/battle/sessions/{sid}/entities/{player_id}/player-card-mode",
            json={"physicalCards": False},
        )
        self.assertEqual(disabled.status_code, 400)
        self.assertIn("requires a deck reset confirmation", disabled.json()["detail"])

        reset = self.client.post(
            f"/api/battle/sessions/{sid}/entities/{player_id}/player-card-mode",
            json={"physicalCards": False, "deckReset": True},
        )
        self.assertEqual(reset.status_code, 200)
        player_payload = next(enemy for enemy in reset.json()["enemies"] if enemy["instance_id"] == player_id)
        self.assertFalse(player_payload["physical_cards"])
        self.assertEqual(player_payload["wound_counts"], {"hand": 0, "discard": 0, "draw_pile": 3, "total": 3})

    def test_player_wound_ko_is_reported_in_snapshot_and_clears_after_hand_wounds(self) -> None:
        sid = self.client.post("/api/battle/sessions").json()["sid"]
        player_snapshot = self.client.post(
            f"/api/battle/sessions/{sid}/players",
            json={"name": "Mira", "power": 1},
        ).json()
        player_id = player_snapshot["selectedId"]
        session = self.context.load_session(sid)
        player = session.state.enemies[player_id]
        player.deck_state.hand = [WOUND_CARD_ID]
        player.deck_state.discard_pile = []
        player.deck_state.draw_pile = ["hf_martial_1_success"]
        session.active_turn_id = player_id
        session.autosave()

        drawn = self.client.post(f"/api/battle/sessions/{sid}/turn/draw")
        self.assertEqual(drawn.status_code, 200)
        player_payload = next(enemy for enemy in drawn.json()["enemies"] if enemy["instance_id"] == player_id)
        self.assertTrue(player_payload["is_ko"])
        self.assertTrue(player_payload["is_down"])

        discarded = self.client.post(f"/api/battle/sessions/{sid}/entities/{player_id}/wounds/discard")
        self.assertEqual(discarded.status_code, 200)
        player_payload = next(enemy for enemy in discarded.json()["enemies"] if enemy["instance_id"] == player_id)
        self.assertFalse(player_payload["is_ko"])
        self.assertFalse(player_payload["is_down"])

    def test_heal_endpoint_clamps_player_to_toughness_max(self) -> None:
        sid = self.client.post("/api/battle/sessions").json()["sid"]
        self.client.post(
            f"/api/battle/sessions/{sid}/players",
            json={"name": "Mira", "toughness": 3, "armor": 0, "magicArmor": 0, "power": 0, "movement": 6},
        )
        self.client.post(
            f"/api/battle/sessions/{sid}/attack",
            json={"damage": 1, "burn": False, "poison": False, "slow": False, "paralyze": False, "modifiers": []},
        )

        payload = self.client.post(
            f"/api/battle/sessions/{sid}/heal",
            json={"toughness": 3, "armor": 0, "magicArmor": 0, "C_HOBGOBLIN": 0},
        ).json()

        player = next(enemy for enemy in payload["enemies"] if enemy["template_id"] == "player")
        self.assertEqual(player["toughness_current"], 3)
        self.assertEqual(player["toughness_max"], 3)

        payload = self.client.post(
            f"/api/battle/sessions/{sid}/heal",
            json={"toughness": 3, "armor": 0, "magicArmor": 0, "C_HOBGOBLIN": 0},
        ).json()

        player = next(enemy for enemy in payload["enemies"] if enemy["template_id"] == "player")
        self.assertEqual(player["toughness_current"], 3)

    def test_heal_endpoint_can_add_player_temporary_toughness(self) -> None:
        sid = self.client.post("/api/battle/sessions").json()["sid"]
        self.client.post(
            f"/api/battle/sessions/{sid}/players",
            json={"name": "Mira", "toughness": 4, "armor": 0, "magicArmor": 0, "power": 0, "movement": 6},
        )
        self.client.post(
            f"/api/battle/sessions/{sid}/attack",
            json={"damage": 1, "burn": False, "poison": False, "slow": False, "paralyze": False, "modifiers": []},
        )

        payload = self.client.post(
            f"/api/battle/sessions/{sid}/heal",
            json={"toughness": 5, "temporaryToughness": 2, "armor": 0, "magicArmor": 0, "guard": 0},
        ).json()

        player = next(enemy for enemy in payload["enemies"] if enemy["template_id"] == "player")
        self.assertEqual(player["toughness_current"], 6)
        self.assertEqual(player["toughness_max"], 4)

    def test_quick_attack_endpoint_uses_active_draw_and_supports_undo(self) -> None:
        sid = self.client.post("/api/battle/sessions").json()["sid"]
        attacker_snapshot = self.client.post(f"/api/battle/sessions/{sid}/enemies", json={"templateId": "C_WOLF"}).json()
        attacker_id = attacker_snapshot["selectedId"]
        target_snapshot = self.client.post(f"/api/battle/sessions/{sid}/enemies", json={"templateId": "C_GOBLIN"}).json()
        target_id = target_snapshot["selectedId"]
        session = self.context.load_session(sid)
        attacker = session.state.enemies[attacker_id]
        target = session.state.enemies[target_id]
        target.toughness_current = 10
        target.toughness_max = 10
        target.guard_current = 0
        target.armor_current = 0
        target.armor_max = 0
        self.context.card_index["test_stab_5"] = Card(
            id="test_stab_5",
            title="Test Stab",
            effects=(Effect(type="attack", amount=5, modifiers=("stab",)),),
        )
        attacker.deck_state.hand = ["test_stab_5"]
        session.active_turn_id = attacker_id
        session.turn_in_progress = True
        session.select(target_id)
        session.autosave()

        quick_response = self.client.post(f"/api/battle/sessions/{sid}/turn/quick-attack")

        self.assertEqual(quick_response.status_code, 200)
        payload = quick_response.json()
        attacked = next(enemy for enemy in payload["enemies"] if enemy["instance_id"] == target_id)
        attacker_payload = next(enemy for enemy in payload["enemies"] if enemy["instance_id"] == attacker_id)
        self.assertEqual(attacked["toughness_current"], 5)
        self.assertTrue(attacker_payload["quick_attack_used"])
        self.assertTrue(payload["canUndo"])
        self.assertEqual(payload["quickAttack"]["attackerId"], attacker_id)
        self.assertIn("Quick Attack", payload["quickAttackNotice"])

        repeat_response = self.client.post(f"/api/battle/sessions/{sid}/turn/quick-attack")
        self.assertEqual(repeat_response.status_code, 400)
        self.assertIn("already been used", repeat_response.json()["detail"])

        undo_response = self.client.post(f"/api/battle/sessions/{sid}/undo")
        restored = next(enemy for enemy in undo_response.json()["enemies"] if enemy["instance_id"] == target_id)
        restored_attacker = next(enemy for enemy in undo_response.json()["enemies"] if enemy["instance_id"] == attacker_id)
        self.assertEqual(restored["toughness_current"], 10)
        self.assertFalse(restored_attacker["quick_attack_used"])

    def test_undo_restores_previous_mutation_state(self) -> None:
        sid = self.client.post("/api/battle/sessions").json()["sid"]
        enemy_snapshot = self.client.post(f"/api/battle/sessions/{sid}/enemies", json={"templateId": "C_WOLF"}).json()
        entity_id = enemy_snapshot["selectedId"]
        before = next(enemy for enemy in enemy_snapshot["enemies"] if enemy["instance_id"] == entity_id)

        attack_response = self.client.post(
            f"/api/battle/sessions/{sid}/attack",
            json={"damage": 2, "burn": True, "poison": False, "slow": False, "paralyze": False, "modifiers": []},
        )
        self.assertEqual(attack_response.status_code, 200)
        self.assertTrue(attack_response.json()["canUndo"])

        undo_response = self.client.post(f"/api/battle/sessions/{sid}/undo")
        self.assertEqual(undo_response.status_code, 200)
        self.assertTrue(undo_response.json()["canRedo"])
        restored = next(enemy for enemy in undo_response.json()["enemies"] if enemy["instance_id"] == entity_id)
        self.assertEqual(restored["toughness_current"], before["toughness_current"])
        self.assertNotIn("burn", restored["statuses"])
        self.assertEqual(undo_response.json()["combatLog"], enemy_snapshot["combatLog"])

        redo_response = self.client.post(f"/api/battle/sessions/{sid}/redo")
        self.assertEqual(redo_response.status_code, 200)
        redone = next(enemy for enemy in redo_response.json()["enemies"] if enemy["instance_id"] == entity_id)
        attacked = next(enemy for enemy in attack_response.json()["enemies"] if enemy["instance_id"] == entity_id)
        self.assertEqual(redone["toughness_current"], attacked["toughness_current"])
        self.assertIn("burn", redone["statuses"])
        self.assertEqual(redo_response.json()["combatLog"], attack_response.json()["combatLog"])
        self.assertFalse(redo_response.json()["canRedo"])

    def test_redo_persists_and_new_mutation_clears_redo_history(self) -> None:
        sid = self.client.post("/api/battle/sessions").json()["sid"]
        add_response = self.client.post(f"/api/battle/sessions/{sid}/enemies", json={"templateId": "C_WOLF"}).json()
        entity_id = add_response["selectedId"]
        attacked_response = self.client.post(
            f"/api/battle/sessions/{sid}/attack",
            json={"damage": 4, "burn": False, "poison": False, "slow": False, "paralyze": False, "modifiers": []},
        ).json()

        undo_response = self.client.post(f"/api/battle/sessions/{sid}/undo").json()
        self.assertEqual(undo_response["redoDepth"], 1)
        reloaded = self.client.get(f"/api/battle/sessions/{sid}").json()
        self.assertEqual(reloaded["redoDepth"], 1)

        redo_response = self.client.post(f"/api/battle/sessions/{sid}/redo").json()
        redone = next(enemy for enemy in redo_response["enemies"] if enemy["instance_id"] == entity_id)
        attacked = next(enemy for enemy in attacked_response["enemies"] if enemy["instance_id"] == entity_id)
        self.assertEqual(redone["toughness_current"], attacked["toughness_current"])
        self.assertEqual(redo_response["redoDepth"], 0)

        self.client.post(f"/api/battle/sessions/{sid}/undo")
        diverged_response = self.client.post(f"/api/battle/sessions/{sid}/players", json={}).json()
        self.assertFalse(diverged_response["canRedo"])
        self.assertEqual(diverged_response["redoDepth"], 0)

    def test_undo_is_lifo_and_persists_across_session_reload(self) -> None:
        sid = self.client.post("/api/battle/sessions").json()["sid"]
        add_response = self.client.post(f"/api/battle/sessions/{sid}/enemies", json={"templateId": "C_WOLF"}).json()
        entity_id = add_response["selectedId"]
        attacked_response = self.client.post(
            f"/api/battle/sessions/{sid}/attack",
            json={"damage": 3, "burn": True, "poison": False, "slow": False, "paralyze": False, "modifiers": []},
        ).json()
        healed_response = self.client.post(
            f"/api/battle/sessions/{sid}/heal",
            json={"toughness": 1, "armor": 0, "magicArmor": 0, "C_HOBGOBLIN": 0},
        ).json()

        self.assertEqual(healed_response["undoDepth"], 3)
        reloaded = self.client.get(f"/api/battle/sessions/{sid}").json()
        self.assertEqual(reloaded["undoDepth"], 3)

        undo_heal = self.client.post(f"/api/battle/sessions/{sid}/undo").json()
        self.assertEqual(undo_heal["undoDepth"], 2)
        after_undo_heal = next(enemy for enemy in undo_heal["enemies"] if enemy["instance_id"] == entity_id)
        attacked_enemy = next(enemy for enemy in attacked_response["enemies"] if enemy["instance_id"] == entity_id)
        self.assertEqual(after_undo_heal["toughness_current"], attacked_enemy["toughness_current"])

        undo_attack = self.client.post(f"/api/battle/sessions/{sid}/undo").json()
        self.assertEqual(undo_attack["undoDepth"], 1)
        after_undo_attack = next(enemy for enemy in undo_attack["enemies"] if enemy["instance_id"] == entity_id)
        added_enemy = next(enemy for enemy in add_response["enemies"] if enemy["instance_id"] == entity_id)
        self.assertEqual(after_undo_attack["toughness_current"], added_enemy["toughness_current"])
        self.assertNotIn("burn", after_undo_attack["statuses"])

    def test_select_is_not_undoable_and_empty_undo_reports_error(self) -> None:
        sid = self.client.post("/api/battle/sessions").json()["sid"]
        empty_response = self.client.post(f"/api/battle/sessions/{sid}/undo")
        self.assertEqual(empty_response.status_code, 400)
        self.assertIn("Nothing to undo", empty_response.json()["detail"])
        empty_redo_response = self.client.post(f"/api/battle/sessions/{sid}/redo")
        self.assertEqual(empty_redo_response.status_code, 400)
        self.assertIn("Nothing to redo", empty_redo_response.json()["detail"])

        first = self.client.post(f"/api/battle/sessions/{sid}/enemies", json={"templateId": "C_GOBLIN"}).json()
        first_id = first["selectedId"]
        second = self.client.post(f"/api/battle/sessions/{sid}/enemies", json={"templateId": "C_WOLF"}).json()
        second_id = second["selectedId"]
        depth_before_select = second["undoDepth"]

        select_response = self.client.post(f"/api/battle/sessions/{sid}/select", json={"instanceId": first_id})
        self.assertEqual(select_response.status_code, 200)
        self.assertEqual(select_response.json()["undoDepth"], depth_before_select)

        undo_response = self.client.post(f"/api/battle/sessions/{sid}/undo").json()
        self.assertEqual(undo_response["undoDepth"], depth_before_select - 1)
        self.assertFalse(any(enemy["instance_id"] == second_id for enemy in undo_response["enemies"]))

    def test_undo_stack_trims_to_twenty_entries(self) -> None:
        sid = self.client.post("/api/battle/sessions").json()["sid"]

        latest = None
        for _ in range(25):
            latest = self.client.post(f"/api/battle/sessions/{sid}/players", json={}).json()

        self.assertEqual(latest["undoDepth"], 20)

    def test_manual_load_resets_undo_history(self) -> None:
        sid = self.client.post("/api/battle/sessions").json()["sid"]
        first = self.client.post(f"/api/battle/sessions/{sid}/enemies", json={"templateId": "C_GOBLIN"}).json()
        self.assertTrue(first["canUndo"])

        save_response = self.client.post(f"/api/battle/sessions/{sid}/saves", json={"name": "undo-reset"})
        saves = self.client.get(f"/api/battle/sessions/{sid}/saves").json()["saves"]
        self.assertTrue(save_response.json()["canUndo"])

        self.client.post(f"/api/battle/sessions/{sid}/enemies", json={"templateId": "C_WOLF"})
        load_response = self.client.post(
            f"/api/battle/sessions/{sid}/load",
            json={"filename": saves[0]["filename"]},
        )

        self.assertEqual(load_response.status_code, 200)
        self.assertFalse(load_response.json()["canUndo"])
        self.assertEqual(load_response.json()["undoDepth"], 0)
        self.assertFalse(load_response.json()["canRedo"])
        self.assertEqual(load_response.json()["redoDepth"], 0)

    def test_player_and_custom_enemy_endpoints_remain_usable(self) -> None:
        sid = self.client.post("/api/battle/sessions").json()["sid"]

        player_response = self.client.post(
            f"/api/battle/sessions/{sid}/players",
            json={"name": "Aldric", "toughness": 20, "armor": 2, "magicArmor": 0, "power": 1, "movement": 5},
        )
        self.assertEqual(player_response.status_code, 200)
        player = next(e for e in player_response.json()["enemies"] if e["template_id"] == "player")
        self.assertEqual(player["name"], "Aldric")
        self.assertEqual(player["toughness_max"], 20)
        self.assertEqual(player["armor_max"], 2)
        self.assertEqual(player["movement"], 5)

        custom_response = self.client.post(
            f"/api/battle/sessions/{sid}/enemies",
            json={
                "custom": {
                    "name": "Shade",
                    "toughness": 7,
                    "armor": 1,
                    "magicArmor": 0,
                    "power": 2,
                    "movement": 4,
                    "coreDeckId": "basic",
                }
            },
        )
        self.assertEqual(custom_response.status_code, 200)
        self.assertTrue(any(enemy["name"] == "Shade" for enemy in custom_response.json()["enemies"]))

    def test_position_endpoint_validates_map_state(self) -> None:
        sid = self.client.post("/api/battle/sessions").json()["sid"]
        first = self.client.post(f"/api/battle/sessions/{sid}/enemies", json={"templateId": "C_GOBLIN"}).json()
        first_id = first["selectedId"]
        second = self.client.post(f"/api/battle/sessions/{sid}/enemies", json={"templateId": "C_WOLF"}).json()
        second_id = second["selectedId"]

        move_response = self.client.post(
            f"/api/battle/sessions/{sid}/entities/{first_id}/position",
            json={"x": 0, "y": 0},
        )
        self.assertEqual(move_response.status_code, 200)
        moved = next(enemy for enemy in move_response.json()["enemies"] if enemy["instance_id"] == first_id)
        self.assertEqual((moved["grid_x"], moved["grid_y"]), (0, 0))
        self.assertEqual(move_response.json()["selectedId"], first_id)

        occupied_response = self.client.post(
            f"/api/battle/sessions/{sid}/entities/{second_id}/position",
            json={"x": 0, "y": 0},
        )
        self.assertEqual(occupied_response.status_code, 400)
        self.assertIn("occupied", occupied_response.json()["detail"])

        # Down units don't block â€” second can stack on the cell where first is downed
        self.client.post(f"/api/battle/sessions/{sid}/select", json={"instanceId": first_id})
        down_response = self.client.post(
            f"/api/battle/sessions/{sid}/attack",
            json={"damage": 999},
        )
        self.assertEqual(down_response.status_code, 200)
        down_first = next(enemy for enemy in down_response.json()["enemies"] if enemy["instance_id"] == first_id)
        self.assertTrue(down_first["is_down"])

        passable_response = self.client.post(
            f"/api/battle/sessions/{sid}/entities/{second_id}/position",
            json={"x": 0, "y": 0},
        )
        self.assertEqual(passable_response.status_code, 200)
        stacked_second = next(enemy for enemy in passable_response.json()["enemies"] if enemy["instance_id"] == second_id)
        self.assertEqual((stacked_second["grid_x"], stacked_second["grid_y"]), (0, 0))

    def test_pending_new_round_flag_and_round_start_endpoint(self) -> None:
        sid = self.client.post("/api/battle/sessions").json()["sid"]
        self.client.post(f"/api/battle/sessions/{sid}/enemies", json={"templateId": "C_GOBLIN"})
        first_id = self.client.post(f"/api/battle/sessions/{sid}/enemies", json={"templateId": "C_WOLF"}).json()["selectedId"]
        self.client.post(f"/api/battle/sessions/{sid}/select", json={"instanceId": self.client.get(f"/api/battle/sessions/{sid}").json()["order"][0]})

        first_next = self.client.post(f"/api/battle/sessions/{sid}/turn/next").json()
        self.assertFalse(first_next["pendingNewRound"])
        self.assertIsNotNone(first_next["activeTurnId"])

        second_next = self.client.post(f"/api/battle/sessions/{sid}/turn/next").json()
        self.assertTrue(second_next["pendingNewRound"])
        self.assertIsNone(second_next["activeTurnId"])
        self.assertEqual(second_next["round"], 1)

        # pendingNewRound persists across reload
        reloaded = self.client.get(f"/api/battle/sessions/{sid}").json()
        self.assertTrue(reloaded["pendingNewRound"])

        start = self.client.post(f"/api/battle/sessions/{sid}/round/start").json()
        self.assertFalse(start["pendingNewRound"])
        self.assertEqual(start["round"], 2)
        self.assertIsNotNone(start["activeTurnId"])

    def test_end_combat_endpoint_resets_combat_state(self) -> None:
        sid = self.client.post("/api/battle/sessions").json()["sid"]
        self.client.post(f"/api/battle/sessions/{sid}/enemies", json={"templateId": "C_GOBLIN"})
        self.client.post(f"/api/battle/sessions/{sid}/encounter/start")
        second_round = self.client.post(f"/api/battle/sessions/{sid}/turn/next").json()
        if second_round["pendingNewRound"]:
            self.client.post(f"/api/battle/sessions/{sid}/round/start")

        response = self.client.post(f"/api/battle/sessions/{sid}/encounter/end")

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertFalse(payload["encounterStarted"])
        self.assertIsNone(payload["activeTurnId"])
        self.assertFalse(payload["turnInProgress"])
        self.assertFalse(payload["pendingNewRound"])
        self.assertEqual(payload["round"], 1)
        self.assertIn("Combat ended.", payload["combatLog"])

    def test_add_player_with_stats_endpoint(self) -> None:
        sid = self.client.post("/api/battle/sessions").json()["sid"]

        response = self.client.post(
            f"/api/battle/sessions/{sid}/players",
            json={"name": "Aldric", "toughness": 15, "armor": 2, "magicArmor": 1, "power": 0, "movement": 6},
        )
        self.assertEqual(response.status_code, 200)
        player = next(e for e in response.json()["enemies"] if e["template_id"] == "player")
        self.assertEqual(player["name"], "Aldric")
        self.assertEqual(player["toughness_max"], 15)
        self.assertEqual(player["armor_max"], 2)
        self.assertEqual(player["magic_armor_max"], 1)
        self.assertEqual(player["movement"], 6)

        # default name when omitted
        default_response = self.client.post(f"/api/battle/sessions/{sid}/players", json={})
        self.assertEqual(default_response.status_code, 200)
        default_player = next(e for e in default_response.json()["enemies"] if "Player" in e["name"])
        self.assertIn("Player", default_player["name"])
        self.assertEqual(default_player["toughness_max"], 4)
        self.assertEqual(default_player["armor_max"], 1)
        self.assertEqual(default_player["guard_base"], 1)
        self.assertEqual(default_player["power_base"], 4)
        self.assertEqual(default_player["initiative_modifier"], 2)

    def test_add_player_can_select_player_deck(self) -> None:
        sid = self.client.post("/api/battle/sessions").json()["sid"]

        response = self.client.post(
            f"/api/battle/sessions/{sid}/players",
            json={"name": "Merlin", "playerDeckId": "human_wizzard_lvl1"},
        )

        self.assertEqual(response.status_code, 200)
        player = next(e for e in response.json()["enemies"] if e["template_id"] == "player")
        self.assertEqual(player["name"], "Merlin")
        self.assertEqual(player["core_deck_id"], "human_wizzard_lvl1")

    def test_player_draw_exact_endpoint_allows_skill_draw_after_power_draw(self) -> None:
        sid = self.client.post("/api/battle/sessions").json()["sid"]
        added = self.client.post(f"/api/battle/sessions/{sid}/players", json={"name": "Mira"}).json()
        player_id = added["selectedId"]

        first = self.client.post(f"/api/battle/sessions/{sid}/turn/draw")
        self.assertEqual(first.status_code, 200)
        first_payload = first.json()
        self.assertEqual(first_payload["activeTurnId"], player_id)
        self.assertTrue(first_payload["turnInProgress"])

        second_power = self.client.post(f"/api/battle/sessions/{sid}/turn/draw")
        self.assertEqual(second_power.status_code, 400)

        second = self.client.post(f"/api/battle/sessions/{sid}/turn/draw-exact", json={"count": 1})
        self.assertEqual(second.status_code, 200)
        player = next(e for e in second.json()["enemies"] if e["instance_id"] == player_id)
        self.assertEqual(len(player["current_draw_groups"]), 2)

    def test_guard_action_endpoint_adds_guard(self) -> None:
        sid = self.client.post("/api/battle/sessions").json()["sid"]
        added = self.client.post(f"/api/battle/sessions/{sid}/players", json={"name": "Mira"}).json()
        player_id = added["selectedId"]

        response = self.client.post(f"/api/battle/sessions/{sid}/action/guard", json={"x": 3})

        self.assertEqual(response.status_code, 200)
        player = next(e for e in response.json()["enemies"] if e["instance_id"] == player_id)
        self.assertEqual(player["guard_current"], 3)
        self.assertEqual(player["actions_used"], 1)
        self.assertIn("Mira guards: +3 guard.", response.json()["combatLog"])

    def test_hitdraw_endpoint_returns_result_after_draw_of_power(self) -> None:
        sid = self.client.post("/api/battle/sessions").json()["sid"]
        added = self.client.post(f"/api/battle/sessions/{sid}/players", json={"name": "Mira", "power": 1}).json()
        player_id = added["selectedId"]
        session = self.context.load_session(sid)
        player = session.state.enemies[player_id]
        player.deck_state.draw_pile = ["hf_martial_success_2"]
        player.deck_state.discard_pile = []
        player.deck_state.hand = []
        session.active_turn_id = player_id
        session.autosave()

        early = self.client.post(f"/api/battle/sessions/{sid}/action/hitdraw")
        self.assertEqual(early.status_code, 400)
        self.assertIn("Draw of Power", early.json()["detail"])

        session = self.context.load_session(sid)
        session.active_turn_id = None
        session.autosave()

        self.assertEqual(self.client.post(f"/api/battle/sessions/{sid}/turn/draw").status_code, 200)
        session = self.context.load_session(sid)
        player = session.state.enemies[player_id]
        player.deck_state.draw_pile = ["hf_martial_success_3", "hf_martial_fate_1", WOUND_CARD_ID]
        player.deck_state.discard_pile = []
        session.autosave()

        response = self.client.post(f"/api/battle/sessions/{sid}/action/hitdraw")

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["hitDraw"]["drawnText"], ["Success", "Fate", "Fail"])
        self.assertEqual(
            payload["hitDraw"]["drawnCards"],
            [
                {"label": "Success", "detail": "Martial 3 energy"},
                {"label": "Fate", "detail": "Martial 1 energy"},
                {"label": "Fail", "detail": "Wound"},
            ],
        )
        self.assertEqual(payload["hitDraw"]["summary"]["outcomes"], {"success": 1, "fate": 1, "fail": 1})
        player_payload = next(e for e in payload["enemies"] if e["instance_id"] == player_id)
        self.assertEqual(player_payload["actions_used"], 1)
        self.assertEqual(player_payload["current_draw_text"], ["Martial 2 energy success"])

    def test_hitdraw_endpoint_rejects_physical_players(self) -> None:
        sid = self.client.post("/api/battle/sessions").json()["sid"]
        self.client.post(
            f"/api/battle/sessions/{sid}/players",
            json={"name": "Mira", "physicalCards": True},
        )

        response = self.client.post(f"/api/battle/sessions/{sid}/action/hitdraw")

        self.assertEqual(response.status_code, 400)
        self.assertIn("Physical-card players", response.json()["detail"])

    def test_roll_initiative_endpoint(self) -> None:
        sid = self.client.post("/api/battle/sessions").json()["sid"]
        self.client.post(f"/api/battle/sessions/{sid}/enemies", json={"templateId": "C_GOBLIN"})
        self.client.post(f"/api/battle/sessions/{sid}/enemies", json={"templateId": "C_WOLF"})

        response = self.client.post(f"/api/battle/sessions/{sid}/initiative/roll", json={"modes": {}})

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["initiativeRolledRound"], 1)

        # All enemies should have initiative_total set
        for enemy in payload["enemies"]:
            self.assertIsNotNone(enemy["initiative_total"])
            self.assertIsNotNone(enemy["initiative_roll"])

    def test_roll_initiative_blocked_when_encounter_active(self) -> None:
        sid = self.client.post("/api/battle/sessions").json()["sid"]
        self.client.post(f"/api/battle/sessions/{sid}/enemies", json={"templateId": "C_GOBLIN"})
        self.client.post(f"/api/battle/sessions/{sid}/encounter/start")

        response = self.client.post(f"/api/battle/sessions/{sid}/initiative/roll", json={"modes": {}})
        self.assertEqual(response.status_code, 400)
        self.assertIn("Cannot roll initiative", response.json()["detail"])

    def test_combat_sim_single_endpoint_returns_timeline_and_stats(self) -> None:
        response = self.client.post(
            "/api/combat-sim/simulate",
            json={
                "teamA": [{"templateId": "C_GOBLIN", "count": 1}],
                "teamB": [{"templateId": "C_WOLF", "count": 1}],
                "strategyA": "highest_toughness",
                "strategyB": "highest_toughness",
                "seed": 123,
                "runs": 1,
            },
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["mode"], "single")
        result = payload["result"]
        self.assertEqual(result["seed"], 123)
        self.assertIn(result["winner"], {"A", "B", "draw"})
        self.assertGreaterEqual(result["rounds"], 1)
        self.assertGreaterEqual(len(result["initialUnits"]), 2)
        self.assertGreaterEqual(len(result["finalUnits"]), 2)
        self.assertGreaterEqual(len(result["timeline"]), 1)
        self.assertIn("initiativeText", result["initialUnits"][0])
        self.assertIn("teamTotals", result)
        self.assertIn("coverageSummary", result)

    def test_combat_sim_endpoint_accepts_entry_overrides_without_mutating_metadata(self) -> None:
        before_meta = self.client.get("/api/battle/meta").json()
        before_goblin = next(item for item in before_meta["enemyTemplates"] if item["id"] == "C_GOBLIN")
        source_toughness = before_goblin["simStats"]["toughness"]["value"] or before_goblin["simStats"]["toughness"]["min"]
        source_alertness = before_goblin["skills"]["alertness"]
        action_result = before_goblin["simActions"][0]["result"]

        response = self.client.post(
            "/api/combat-sim/simulate",
            json={
                "teamA": [
                    {
                        "templateId": "C_GOBLIN",
                        "count": 1,
                        "overrides": {
                            "statOverrides": {"toughness": source_toughness + 5, "armor": 3},
                            "skillOverrides": {"alertness": source_alertness + 4},
                            "actionOverrides": {action_result: "Test Strike - Attack 9"},
                        },
                    }
                ],
                "teamB": [{"templateId": "C_WOLF", "count": 1}],
                "strategy": "highest_toughness",
                "seed": 500,
                "runs": 1,
            },
        )

        self.assertEqual(response.status_code, 200)
        result = response.json()["result"]
        goblin = next(unit for unit in result["initialUnits"] if unit["templateId"] == "C_GOBLIN")
        self.assertEqual(goblin["toughnessMax"], source_toughness + 5)
        self.assertEqual(goblin["armorMax"], 3)
        self.assertEqual(goblin["initiativeModifier"], source_alertness + 4)

        after_meta = self.client.get("/api/battle/meta").json()
        after_goblin = next(item for item in after_meta["enemyTemplates"] if item["id"] == "C_GOBLIN")
        self.assertEqual(after_goblin["simStats"]["toughness"], before_goblin["simStats"]["toughness"])
        self.assertEqual(after_goblin["skills"], before_goblin["skills"])
        self.assertEqual(after_goblin["simActions"], before_goblin["simActions"])

    def test_save_creature_template_overrides_writes_workbook_and_reloads_metadata(self) -> None:
        client, context, workbook_path = self._client_with_temp_workbook()
        sid = client.post("/api/battle/sessions").json()["sid"]
        added = client.post(f"/api/battle/sessions/{sid}/enemies", json={"templateId": "C_GOBLIN"}).json()
        entity_id = added["selectedId"]
        before_enemy = next(enemy for enemy in added["enemies"] if enemy["instance_id"] == entity_id)

        response = client.post(
            "/api/battle/creature-templates/C_GOBLIN/save-overrides",
            json={
                "statOverrides": {"toughness": 14, "armor": 3},
                "skillOverrides": {"alertness": 7},
                "actionOverrides": {"A1": "Saved Strike - Attack 8 pierce 1"},
                "infoOverrides": {"playtestStatus": "Retest_Needed"},
            },
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["backupFilename"].startswith("denizens_creature_database__"))
        self.assertTrue((context.creature_workbook_backup_dir / payload["backupFilename"]).exists())
        goblin_meta = next(item for item in payload["metadata"]["enemyTemplates"] if item["id"] == "C_GOBLIN")
        self.assertEqual(goblin_meta["simStats"]["toughness"]["value"], 14)
        self.assertEqual(goblin_meta["simStats"]["armor"]["value"], 3)
        self.assertEqual(goblin_meta["skills"]["alertness"], 7)
        self.assertEqual(goblin_meta["simStats"]["initiativeModifier"], 7)
        self.assertEqual(goblin_meta["playtestStatus"], "Retest_Needed")
        self.assertEqual(next(action for action in goblin_meta["simActions"] if action["result"] == "A1")["text"], "Saved Strike - Attack 8 pierce 1")
        self.assertEqual(context.enemy_templates["C_GOBLIN"].initiative_modifier, 7)

        values = self._workbook_values(workbook_path, "C_GOBLIN", ["Toughness", "Armor", "Alertness", "A1", "Playtest_Status"])
        self.assertEqual(values["Toughness"], 14)
        self.assertEqual(values["Armor"], 3)
        self.assertEqual(values["Alertness"], 7)
        self.assertEqual(values["A1"], "Saved Strike - Attack 8 pierce 1")
        self.assertEqual(values["Playtest_Status"], "Retest_Needed")

        after_session = client.get(f"/api/battle/sessions/{sid}").json()
        after_enemy = next(enemy for enemy in after_session["enemies"] if enemy["instance_id"] == entity_id)
        self.assertEqual(after_enemy["toughness_max"], before_enemy["toughness_max"])

    def test_save_creature_template_overrides_rejects_invalid_payloads(self) -> None:
        client, _context, _workbook_path = self._client_with_temp_workbook()

        bad_init = client.post(
            "/api/battle/creature-templates/C_GOBLIN/save-overrides",
            json={"statOverrides": {"initiativeModifier": 4}},
        )
        self.assertEqual(bad_init.status_code, 400)
        self.assertIn("Alertness", bad_init.json()["detail"])

        bad_action = client.post(
            "/api/battle/creature-templates/C_GOBLIN/save-overrides",
            json={"actionOverrides": {"A1": "Broken - Attack target"}},
        )
        self.assertEqual(bad_action.status_code, 400)
        self.assertIn("not simulatable", bad_action.json()["detail"])

        bad_template = client.post(
            "/api/battle/creature-templates/NOPE/save-overrides",
            json={"statOverrides": {"toughness": 10}},
        )
        self.assertEqual(bad_template.status_code, 400)
        self.assertIn("Unknown template", bad_template.json()["detail"])

    def test_combat_sim_batch_endpoint_returns_aggregate_and_last_combat(self) -> None:
        response = self.client.post(
            "/api/combat-sim/simulate",
            json={
                "teamA": [{"templateId": "C_GOBLIN", "count": 2}],
                "teamB": [{"templateId": "C_WOLF", "count": 1}],
                "strategy": "highest_toughness",
                "seed": 200,
                "runs": 4,
            },
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["mode"], "batch")
        result = payload["result"]
        self.assertEqual(result["seed"], 200)
        self.assertEqual(result["runs"], 4)
        self.assertEqual(result["lastCombat"]["seed"], 203)
        self.assertEqual(sum(result["summary"]["wins"].values()), 4)
        self.assertIn("avgRounds", result["summary"])
        self.assertIn("teamAverages", result["summary"])

    def test_combat_sim_batch_precision_endpoint_uses_cap_and_reports_ci(self) -> None:
        response = self.client.post(
            "/api/combat-sim/simulate",
            json={
                "teamA": [{"templateId": "C_GOBLIN", "count": 1}],
                "teamB": [{"templateId": "C_WOLF", "count": 1}],
                "strategy": "highest_toughness",
                "seed": 300,
                "runs": 10,
                "precisionTargetPercent": 50,
            },
        )

        self.assertEqual(response.status_code, 200)
        result = response.json()["result"]
        self.assertLessEqual(result["runs"], 10)
        precision = result["summary"]["precision"]
        self.assertEqual(precision["targetRerunFluctuation"], 0.5)
        self.assertIn("worstCaseRerunFluctuation95", precision)
        self.assertIn("ciLow", precision["outcomes"]["A"])
        self.assertIn("std", precision["outcomes"]["A"])

    def test_combat_sim_endpoint_rejects_invalid_input(self) -> None:
        bad_template = self.client.post(
            "/api/combat-sim/simulate",
            json={
                "teamA": [{"templateId": "missing", "count": 1}],
                "teamB": [{"templateId": "C_WOLF", "count": 1}],
                "runs": 1,
            },
        )
        self.assertEqual(bad_template.status_code, 400)
        self.assertIn("Unknown template", bad_template.json()["detail"])

        bad_count = self.client.post(
            "/api/combat-sim/simulate",
            json={
                "teamA": [{"templateId": "C_GOBLIN", "count": 0}],
                "teamB": [{"templateId": "C_WOLF", "count": 1}],
                "runs": 1,
            },
        )
        self.assertEqual(bad_count.status_code, 400)
        self.assertIn("count", bad_count.json()["detail"])

        bad_runs = self.client.post(
            "/api/combat-sim/simulate",
            json={
                "teamA": [{"templateId": "C_GOBLIN", "count": 1}],
                "teamB": [{"templateId": "C_WOLF", "count": 1}],
                "runs": 1001,
            },
        )
        self.assertEqual(bad_runs.status_code, 400)
        self.assertIn("runs", bad_runs.json()["detail"])

        bad_override = self.client.post(
            "/api/combat-sim/simulate",
            json={
                "teamA": [{"templateId": "C_GOBLIN", "count": 1, "overrides": {"statOverrides": {"toughness": 0}}}],
                "teamB": [{"templateId": "C_WOLF", "count": 1}],
                "runs": 1,
            },
        )
        self.assertEqual(bad_override.status_code, 400)
        self.assertIn("toughness", bad_override.json()["detail"])

        bad_skill_override = self.client.post(
            "/api/combat-sim/simulate",
            json={
                "teamA": [{"templateId": "C_GOBLIN", "count": 1, "overrides": {"skillOverrides": {"alertness": -1}}}],
                "teamB": [{"templateId": "C_WOLF", "count": 1}],
                "runs": 1,
            },
        )
        self.assertEqual(bad_skill_override.status_code, 400)
        self.assertIn("alertness", bad_skill_override.json()["detail"])


if __name__ == "__main__":
    unittest.main()
